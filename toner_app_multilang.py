#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Sistem za evidenciju tonera
Pravi: Igor Malkočević
"""

import sys
import sqlite3
from translations import T
from PyQt5.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, 
                             QHBoxLayout, QTabWidget, QTableWidget, QTableWidgetItem,
                             QPushButton, QLineEdit, QLabel, QComboBox, QSpinBox,
                             QMessageBox, QDialog, QFormLayout, QTextEdit, QHeaderView,
                             QCheckBox, QFileDialog, QMenuBar, QAction, QMenu)
from PyQt5.QtCore import Qt, QUrl, QTimer, QSettings
from PyQt5.QtGui import QColor, QDesktopServices

class Database:
    def __init__(self, db_name="toneri.db"):
        self.db_name = db_name
        self.init_database()
    
    def get_connection(self):
        return sqlite3.connect(self.db_name)
    
    def init_database(self):
        """Kreira tabele ako ne postoje"""
        conn = self.get_connection()
        cursor = conn.cursor()
        
        # Tabela TONERI
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS toneri (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                model TEXT NOT NULL UNIQUE,
                opis TEXT,
                minimalna_kolicina INTEGER DEFAULT 2,
                trenutno_stanje INTEGER DEFAULT 0,
                driver_link TEXT
            )
        ''')
        
        # Tabela ŠTAMPAČI
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS stampaci (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                model TEXT NOT NULL,
                serijski_broj TEXT UNIQUE,
                status TEXT DEFAULT 'Aktivan',
                napomena TEXT,
                kolicina INTEGER DEFAULT 1,
                driver_link TEXT
            )
        ''')
        
        # Add kolicina column if it doesn't exist (for existing databases)
        try:
            cursor.execute("ALTER TABLE stampaci ADD COLUMN kolicina INTEGER DEFAULT 1")
        except:
            pass
        
        # Add driver_link column if it doesn't exist (for existing databases)
        try:
            cursor.execute("ALTER TABLE stampaci ADD COLUMN driver_link TEXT")
        except:
            pass
        try:
            cursor.execute("ALTER TABLE stampaci ADD COLUMN kolicina INTEGER DEFAULT 1")
        except:
            pass  # Column already exists
        
        # Tabela RADNICI
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS radnici (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                ime TEXT NOT NULL,
                prezime TEXT NOT NULL,
                odeljenje TEXT
            )
        ''')
        
        # Veza: Štampač koristi tonere (Many-to-Many)
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS stampac_toneri (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                stampac_id INTEGER,
                toner_id INTEGER,
                FOREIGN KEY (stampac_id) REFERENCES stampaci(id),
                FOREIGN KEY (toner_id) REFERENCES toneri(id),
                UNIQUE(stampac_id, toner_id)
            )
        ''')
        
        # Veza: Radnik ima štampače (Many-to-Many)
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS radnik_stampaci (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                radnik_id INTEGER,
                stampac_id INTEGER,
                datum_dodeljivanja DATE DEFAULT CURRENT_DATE,
                FOREIGN KEY (radnik_id) REFERENCES radnici(id),
                FOREIGN KEY (stampac_id) REFERENCES stampaci(id),
                UNIQUE(radnik_id, stampac_id)
            )
        ''')
        
        # Istorija narudžbina
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS istorija_narudzbi (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                datum DATE DEFAULT CURRENT_DATE,
                toner_id INTEGER,
                kolicina INTEGER,
                napomena TEXT,
                FOREIGN KEY (toner_id) REFERENCES toneri(id)
            )
        ''')
        
        # Istorija potrošnje
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS istorija_potrosnje (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                datum DATE DEFAULT CURRENT_DATE,
                toner_id INTEGER,
                kolicina INTEGER DEFAULT 1,
                FOREIGN KEY (toner_id) REFERENCES toneri(id)
            )
        ''')
        
        # Backup settings
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS backup_settings (
                id INTEGER PRIMARY KEY CHECK (id = 1),
                enabled INTEGER DEFAULT 0,
                day_of_month INTEGER DEFAULT 1,
                last_backup_date DATE
            )
        ''')
        
        # Inicijalizuj backup settings ako ne postoji
        cursor.execute("INSERT OR IGNORE INTO backup_settings (id) VALUES (1)")
        
        conn.commit()
        conn.close()


class TonerDialog(QDialog):
    """Dialog za dodavanje/editovanje tonera"""
    def __init__(self, parent=None, toner_data=None):
        super().__init__(parent)
        self.toner_data = toner_data
        self.lang = parent.lang if parent and hasattr(parent, 'lang') else 'sr'
        self.setWindowTitle(T.get("dialog_add_toner", self.lang) if not toner_data else T.get("dialog_edit_toner", self.lang))
        self.setMinimumWidth(400)
        self.init_ui()
    
    def init_ui(self):
        layout = QFormLayout()
        
        self.model_input = QLineEdit()
        self.min_kol_input = QSpinBox()
        self.min_kol_input.setMinimum(0)
        self.min_kol_input.setMaximum(999)
        self.min_kol_input.setValue(2)
        self.stanje_input = QSpinBox()
        self.stanje_input.setMinimum(0)
        self.stanje_input.setMaximum(9999)
        
        if self.toner_data:
            self.model_input.setText(self.toner_data[1] if self.toner_data[1] else "")
            self.min_kol_input.setValue(self.toner_data[2] if self.toner_data[2] is not None else 2)
            self.stanje_input.setValue(self.toner_data[3] if self.toner_data[3] is not None else 0)
        
        layout.addRow(T.get("label_toner_model", self.lang), self.model_input)
        layout.addRow(T.get("label_min_qty", self.lang), self.min_kol_input)
        layout.addRow(T.get("label_current_stock", self.lang), self.stanje_input)
        
        # Dugmad
        btn_layout = QHBoxLayout()
        save_btn = QPushButton(T.get("btn_save", self.lang))
        cancel_btn = QPushButton(T.get("btn_cancel", self.lang))
        save_btn.clicked.connect(self.accept)
        cancel_btn.clicked.connect(self.reject)
        btn_layout.addWidget(save_btn)
        btn_layout.addWidget(cancel_btn)
        
        layout.addRow(btn_layout)
        self.setLayout(layout)
    
    def get_data(self):
        return {
            'model': self.model_input.text().strip(),
            'minimalna_kolicina': self.min_kol_input.value(),
            'trenutno_stanje': self.stanje_input.value()
        }


class StampacDialog(QDialog):
    """Dialog za dodavanje/editovanje štampača"""
    def __init__(self, parent=None, db=None, stampac_data=None):
        super().__init__(parent)
        self.db = db
        self.stampac_data = stampac_data
        self.lang = parent.lang if parent and hasattr(parent, 'lang') else 'sr'
        self.setWindowTitle(T.get("dialog_add_printer", self.lang) if not stampac_data else T.get("dialog_edit_printer", self.lang))
        self.setMinimumWidth(500)
        self.setMinimumHeight(500)
        self.init_ui()
    
    def init_ui(self):
        layout = QVBoxLayout()
        form_layout = QFormLayout()
        
        self.model_input = QLineEdit()
        self.kolicina_input = QSpinBox()
        self.kolicina_input.setMinimum(0)
        self.kolicina_input.setMaximum(999)
        self.kolicina_input.setValue(1)
        self.status_combo = QComboBox()
        self.status_combo.addItems([T.get("status_active", self.lang), T.get("status_in_service", self.lang), T.get("status_for_disposal", self.lang)])
        self.napomena_input = QTextEdit()
        self.napomena_input.setMaximumHeight(60)
        self.driver_input = QTextEdit()
        self.driver_input.setMaximumHeight(60)
        
        if self.stampac_data:
            self.model_input.setText(self.stampac_data[1])
            self.status_combo.setCurrentText(self.stampac_data[3])
            self.napomena_input.setText(self.stampac_data[4] or "")
            # Load kolicina if exists (index 5)
            if len(self.stampac_data) > 5 and self.stampac_data[5]:
                self.kolicina_input.setValue(self.stampac_data[5])
            # Load driver_link if exists (index 6)
            if len(self.stampac_data) > 6 and self.stampac_data[6]:
                self.driver_input.setText(self.stampac_data[6] or "")
        
        form_layout.addRow(T.get("label_printer_model", self.lang), self.model_input)
        form_layout.addRow(T.get("label_quantity", self.lang), self.kolicina_input)
        form_layout.addRow(T.get("label_status", self.lang), self.status_combo)
        form_layout.addRow(T.get("label_note", self.lang), self.napomena_input)
        form_layout.addRow(T.get("label_driver_link", self.lang), self.driver_input)
        
        layout.addLayout(form_layout)
        
        # Lista tonera (checkbox)
        toneri_label = QLabel(T.get("label_toners_used", self.lang))
        toneri_label.setStyleSheet("font-weight: bold; margin-top: 10px;")
        layout.addWidget(toneri_label)
        
        # SEARCH za tonere
        self.search_toner_input = QLineEdit()
        self.search_toner_input.setPlaceholderText("🔍 " + T.get("placeholder_search", self.lang) + "...")
        self.search_toner_input.textChanged.connect(self.filter_toneri)
        layout.addWidget(self.search_toner_input)
        
        self.toneri_table = QTableWidget()
        self.toneri_table.setColumnCount(2)
        self.toneri_table.setHorizontalHeaderLabels(["✓", T.get("col_toner_model", self.lang)])
        self.toneri_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.toneri_table.setColumnWidth(0, 30)
        self.toneri_table.setMaximumHeight(200)
        
        # Učitaj sve tonere
        conn = self.db.get_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT id, model FROM toneri ORDER BY model")
        self.all_toneri = cursor.fetchall()  # Sačuvaj kao class variable
        
        # Ako editujemo štampač, učitaj njegove tonere
        selected_toneri = []
        if self.stampac_data:
            cursor.execute("""
                SELECT toner_id FROM stampac_toneri WHERE stampac_id = ?
            """, (self.stampac_data[0],))
            selected_toneri = [row[0] for row in cursor.fetchall()]
        
        conn.close()
        
        self.selected_toneri_ids = selected_toneri  # Sačuvaj za filter
        self.load_toneri_table()  # Inicijalno punjenje tabele
        
        layout.addWidget(self.toneri_table)
        
        # Dugmad
        btn_layout = QHBoxLayout()
        save_btn = QPushButton(T.get("btn_save", self.lang))
        cancel_btn = QPushButton(T.get("btn_cancel", self.lang))
        save_btn.clicked.connect(self.accept)
        cancel_btn.clicked.connect(self.reject)
        btn_layout.addWidget(save_btn)
        btn_layout.addWidget(cancel_btn)
        
        layout.addLayout(btn_layout)
        self.setLayout(layout)
    
    def get_data(self):
        return {
            'model': self.model_input.text().strip(),
            'kolicina': self.kolicina_input.value(),
            'status': self.status_combo.currentText(),
            'napomena': self.napomena_input.toPlainText().strip(),
            'driver_link': self.driver_input.toPlainText().strip()
        }
    
    def load_toneri_table(self):
        """Učitava tabelu tonera"""
        self.toneri_table.setRowCount(len(self.all_toneri))
        self.toner_checkboxes = {}
        
        for i, (toner_id, toner_model) in enumerate(self.all_toneri):
            # Checkbox
            checkbox = QTableWidgetItem()
            checkbox.setFlags(Qt.ItemIsUserCheckable | Qt.ItemIsEnabled)
            checkbox.setCheckState(Qt.Checked if toner_id in self.selected_toneri_ids else Qt.Unchecked)
            self.toneri_table.setItem(i, 0, checkbox)
            
            # Model
            self.toneri_table.setItem(i, 1, QTableWidgetItem(toner_model))
            self.toner_checkboxes[toner_id] = checkbox
    
    def filter_toneri(self):
        """Filtrira tonere prema search tekstu"""
        search_text = self.search_toner_input.text().lower()
        
        for i in range(self.toneri_table.rowCount()):
            model_item = self.toneri_table.item(i, 1)
            if model_item:
                # Prikaži red ako search tekst odgovara ili je search prazan
                matches = search_text in model_item.text().lower() if search_text else True
                self.toneri_table.setRowHidden(i, not matches)
    
    def get_selected_toneri(self):
        """Vraća listu ID-eva označenih tonera"""
        selected = []
        
        for i, (toner_id, _) in enumerate(self.all_toneri):
            checkbox = self.toneri_table.item(i, 0)
            if checkbox and checkbox.checkState() == Qt.Checked:
                selected.append(toner_id)
        
        return selected


class RadnikDialog(QDialog):
    """Dialog za dodavanje/editovanje radnika"""
    def __init__(self, parent=None, db=None, radnik_data=None):
        super().__init__(parent)
        self.db = db
        self.radnik_data = radnik_data
        self.lang = parent.lang if parent and hasattr(parent, 'lang') else 'sr'
        self.setWindowTitle(T.get("dialog_add_employee", self.lang) if not radnik_data else T.get("dialog_edit_employee", self.lang))
        self.setMinimumWidth(500)
        self.setMinimumHeight(500)
        self.init_ui()
    
    def init_ui(self):
        layout = QVBoxLayout()
        form_layout = QFormLayout()
        
        self.ime_input = QLineEdit()
        self.prezime_input = QLineEdit()
        
        if self.radnik_data:
            self.ime_input.setText(self.radnik_data[1])
            self.prezime_input.setText(self.radnik_data[2])
        
        form_layout.addRow(T.get("label_first_name", self.lang) + ":", self.ime_input)
        form_layout.addRow(T.get("label_last_name", self.lang) + ":", self.prezime_input)
        
        layout.addLayout(form_layout)
        
        # Lista štampača (checkbox)
        stampaci_label = QLabel(T.get("label_printers_employee_has", self.lang))
        stampaci_label.setStyleSheet("font-weight: bold; margin-top: 10px;")
        layout.addWidget(stampaci_label)
        
        # SEARCH za štampače
        self.search_stampac_input = QLineEdit()
        self.search_stampac_input.setPlaceholderText("🔍 " + T.get("placeholder_search", self.lang) + "...")
        self.search_stampac_input.textChanged.connect(self.filter_stampaci)
        layout.addWidget(self.search_stampac_input)
        
        self.stampaci_table = QTableWidget()
        self.stampaci_table.setColumnCount(3)
        self.stampaci_table.setHorizontalHeaderLabels(["✓", T.get("col_printer_model", self.lang), T.get("col_status", self.lang)])
        self.stampaci_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.stampaci_table.setColumnWidth(0, 30)
        self.stampaci_table.setColumnWidth(2, 100)
        self.stampaci_table.setMaximumHeight(200)
        
        # Učitaj sve štampače
        conn = self.db.get_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT id, model, status FROM stampaci ORDER BY model")
        self.all_stampaci = cursor.fetchall()  # Sačuvaj kao class variable
        
        # Ako editujemo radnika, učitaj njegove štampače
        selected_stampaci = []
        if self.radnik_data:
            cursor.execute("""
                SELECT stampac_id FROM radnik_stampaci WHERE radnik_id = ?
            """, (self.radnik_data[0],))
            selected_stampaci = [row[0] for row in cursor.fetchall()]
        
        conn.close()
        
        self.selected_stampaci_ids = selected_stampaci  # Sačuvaj za filter
        self.load_stampaci_table()  # Inicijalno punjenje tabele
        
        layout.addWidget(self.stampaci_table)
        
        # Dugmad
        btn_layout = QHBoxLayout()
        save_btn = QPushButton(T.get("btn_save", self.lang))
        cancel_btn = QPushButton(T.get("btn_cancel", self.lang))
        save_btn.clicked.connect(self.accept)
        cancel_btn.clicked.connect(self.reject)
        btn_layout.addWidget(save_btn)
        btn_layout.addWidget(cancel_btn)
        
        layout.addLayout(btn_layout)
        self.setLayout(layout)
    
    def get_data(self):
        return {
            'ime': self.ime_input.text().strip(),
            'prezime': self.prezime_input.text().strip()
        }
    
    def load_stampaci_table(self):
        """Učitava tabelu štampača"""
        self.stampaci_table.setRowCount(len(self.all_stampaci))
        self.stampac_checkboxes = {}
        
        for i, (stampac_id, stampac_model, status) in enumerate(self.all_stampaci):
            # Checkbox
            checkbox = QTableWidgetItem()
            checkbox.setFlags(Qt.ItemIsUserCheckable | Qt.ItemIsEnabled)
            checkbox.setCheckState(Qt.Checked if stampac_id in self.selected_stampaci_ids else Qt.Unchecked)
            self.stampaci_table.setItem(i, 0, checkbox)
            
            # Model
            self.stampaci_table.setItem(i, 1, QTableWidgetItem(stampac_model))
            
            # Status - translate from database value
            status_map = {
                'Aktivan': T.get('status_active', self.lang),
                'Na servisu': T.get('status_in_service', self.lang),
                'Za rashod': T.get('status_for_disposal', self.lang)
            }
            translated_status = status_map.get(status, status)
            status_item = QTableWidgetItem(translated_status)
            
            # Color coding based on original database value
            if status == "Na servisu":
                status_item.setBackground(QColor(255, 255, 200))
            elif status == "Za rashod":
                status_item.setBackground(QColor(255, 150, 150))
            self.stampaci_table.setItem(i, 2, status_item)
            
            self.stampac_checkboxes[stampac_id] = checkbox
    
    def filter_stampaci(self):
        """Filtrira štampače prema search tekstu"""
        search_text = self.search_stampac_input.text().lower()
        
        for i in range(self.stampaci_table.rowCount()):
            model_item = self.stampaci_table.item(i, 1)
            if model_item:
                # Prikaži red ako search tekst odgovara ili je search prazan
                matches = search_text in model_item.text().lower() if search_text else True
                self.stampaci_table.setRowHidden(i, not matches)
    
    def get_selected_stampaci(self):
        """Vraća listu ID-eva označenih štampača"""
        selected = []
        
        for i, (stampac_id, _, _) in enumerate(self.all_stampaci):
            checkbox = self.stampaci_table.item(i, 0)
            if checkbox and checkbox.checkState() == Qt.Checked:
                selected.append(stampac_id)
        
        return selected


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.settings = QSettings('TonerInventory', 'TonerApp')  # Pamćenje postavki
        self.lang = self.load_language_preference()  # Load saved language
        self.search_active = False  # Flag to prevent load from overwriting search highlighting
        self.db = Database()
        self.create_menu_bar()
        self.setWindowTitle(T.get('app_title', self.lang))
        self.setMinimumSize(1200, 700)
        self.init_ui()
        self.load_all_data()
        self.restore_column_widths()  # Učitaj sačuvane širine kolona
        self.cleanup_old_history()  # Čisti istoriju starije od 2 godine
        self.check_auto_backup()  # Proveri da li treba automatski backup
    
    def load_language_preference(self):
        """Load saved language preference from config file"""
        try:
            import json
            import os
            config_file = 'app_config.json'
            if os.path.exists(config_file):
                with open(config_file, 'r', encoding='utf-8') as f:
                    config = json.load(f)
                    return config.get('language', 'sr')
        except:
            pass
        return 'sr'  # Default to Serbian
    
    def save_language_preference(self):
        """Save language preference to config file"""
        try:
            import json
            config = {'language': self.lang}
            with open('app_config.json', 'w', encoding='utf-8') as f:
                json.dump(config, f, ensure_ascii=False, indent=2)
        except Exception as e:
            print(f"Error saving language preference: {e}")
    
    def create_menu_bar(self):
        """Creates menu bar with language selector"""
        menubar = self.menuBar()
        
        # Language menu
        lang_menu = menubar.addMenu(T.get('menu_language', self.lang))
        
        # Serbian action
        sr_action = QAction(T.get('menu_serbian', self.lang), self)
        sr_action.triggered.connect(lambda: self.change_language('sr'))
        lang_menu.addAction(sr_action)
        
        # English action
        en_action = QAction(T.get('menu_english', self.lang), self)
        en_action.triggered.connect(lambda: self.change_language('en'))
        lang_menu.addAction(en_action)
    
    def change_language(self, new_lang):
        """Changes application language"""
        self.lang = new_lang
        self.save_language_preference()  # Save preference
        
        # Update window title
        self.setWindowTitle(T.get('app_title', self.lang))
        
        # Update tab names
        self.tabs.setTabText(0, T.get('tab_toneri', self.lang))
        self.tabs.setTabText(1, T.get('tab_stampaci', self.lang))
        self.tabs.setTabText(2, T.get('tab_radnici', self.lang))
        self.tabs.setTabText(3, T.get('tab_pregled', self.lang))
        self.tabs.setTabText(4, T.get('tab_statistika', self.lang))
        self.tabs.setTabText(5, T.get('tab_istorija', self.lang))
        self.tabs.setTabText(6, T.get('tab_backup', self.lang))
        
        # Update menu
        self.menuBar().clear()
        self.create_menu_bar()
        
        # Show message
        if self.lang == 'sr':
            QMessageBox.information(self, "Jezik promenjen", "Jezik promenjen na srpski.\n\nNeke izmene zahtevaju restart aplikacije.")
        else:
            QMessageBox.information(self, "Language Changed", "Language changed to English.\n\nSome changes require application restart.")
    
    def init_ui(self):
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)
        
        # Search bar
        search_layout = QHBoxLayout()
        search_label = QLabel(T.get("search_label", self.lang))
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText(T.get("search_placeholder", self.lang))
        self.search_input.textChanged.connect(self.search_all)
        search_layout.addWidget(search_label)
        search_layout.addWidget(self.search_input)
        main_layout.addLayout(search_layout)
        
        # Tabs
        self.tabs = QTabWidget()
        
        # Tab 1: TONERI
        self.toneri_tab = QWidget()
        toneri_layout = QVBoxLayout(self.toneri_tab)
        
        toneri_btn_layout = QHBoxLayout()
        self.add_toner_btn = QPushButton(T.get("btn_add_toner", self.lang))
        self.edit_toner_btn = QPushButton(T.get("btn_edit_toner", self.lang))
        self.delete_toner_btn = QPushButton(T.get("btn_delete_toner", self.lang))
        self.potrosnja_btn = QPushButton(T.get("btn_record_consumption", self.lang))
        
        # Label za ukupan zbir tonera
        self.ukupno_tonera_label = QLabel()
        self.ukupno_tonera_label.setStyleSheet("""
            QLabel {
                background-color: #34495E;
                color: white;
                padding: 8px 15px;
                border-radius: 5px;
                font-weight: bold;
                font-size: 11pt;
            }
        """)
        self.update_ukupno_tonera()  # Inicijalno postavi vrednost
        
        self.excel_toneri_btn = QPushButton("📊 " + T.get("btn_excel_export", self.lang))
        self.excel_toneri_btn.setStyleSheet("background-color: #FF9800; color: white; font-weight: bold;")
        self.stampaj_tonere_btn = QPushButton("🖨️ " + T.get("btn_preview_print", self.lang))
        self.stampaj_tonere_btn.setStyleSheet("background-color: #2196F3; color: white; font-weight: bold;")
        self.narudzba_btn = QPushButton(T.get("btn_order_list", self.lang))
        self.narudzba_btn.setStyleSheet("background-color: #4CAF50; color: white; font-weight: bold;")
        
        self.add_toner_btn.clicked.connect(self.add_toner)
        self.edit_toner_btn.clicked.connect(self.edit_toner)
        self.delete_toner_btn.clicked.connect(self.delete_toner)
        self.potrosnja_btn.clicked.connect(self.evidentira_potrosnju)
        self.excel_toneri_btn.clicked.connect(self.export_tonere_excel)
        self.stampaj_tonere_btn.clicked.connect(self.stampaj_tonere)
        self.narudzba_btn.clicked.connect(self.prikazi_narudzbu)
        
        toneri_btn_layout.addWidget(self.add_toner_btn)
        toneri_btn_layout.addWidget(self.edit_toner_btn)
        toneri_btn_layout.addWidget(self.delete_toner_btn)
        toneri_btn_layout.addWidget(self.potrosnja_btn)
        toneri_btn_layout.addWidget(self.ukupno_tonera_label)  # Dodaj label
        toneri_btn_layout.addStretch()
        toneri_btn_layout.addWidget(self.excel_toneri_btn)
        toneri_btn_layout.addWidget(self.stampaj_tonere_btn)
        toneri_btn_layout.addWidget(self.narudzba_btn)
        
        self.toneri_table = QTableWidget()
        self.toneri_table.setColumnCount(4)
        self.toneri_table.setHorizontalHeaderLabels([T.get("col_id", self.lang), T.get("col_model", self.lang), T.get("col_min_qty", self.lang), T.get("col_stock", self.lang)])
        self.toneri_table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.toneri_table.horizontalHeader().setStretchLastSection(True)
        self.toneri_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.toneri_table.itemChanged.connect(self.on_toner_stanje_changed)
        
        toneri_layout.addLayout(toneri_btn_layout)
        toneri_layout.addWidget(self.toneri_table)
        
        # Tab 2: ŠTAMPAČI
        self.stampaci_tab = QWidget()
        stampaci_layout = QVBoxLayout(self.stampaci_tab)
        
        stampaci_btn_layout = QHBoxLayout()
        self.add_stampac_btn = QPushButton(T.get("btn_add_printer", self.lang))
        self.edit_stampac_btn = QPushButton(T.get("btn_edit_printer", self.lang))
        self.delete_stampac_btn = QPushButton(T.get("btn_delete_printer", self.lang))
        
        self.status_filter = QComboBox()
        self.status_filter.addItems([T.get("status_all", self.lang), T.get("status_active", self.lang), T.get("status_in_service", self.lang), T.get("status_for_disposal", self.lang)])
        self.status_filter.currentTextChanged.connect(self.load_stampaci)
        
        # Label za ukupan broj štampača
        self.ukupno_stampaca_label = QLabel()
        self.ukupno_stampaca_label.setStyleSheet("""
            QLabel {
                background-color: #34495E;
                color: white;
                padding: 8px 15px;
                border-radius: 5px;
                font-weight: bold;
                font-size: 11pt;
            }
        """)
        self.update_ukupno_stampaca()  # Inicijalno postavi vrednost
        
        self.excel_stampaci_btn = QPushButton("📊 " + T.get("btn_excel_export", self.lang))
        self.excel_stampaci_btn.setStyleSheet("background-color: #FF9800; color: white; font-weight: bold;")
        self.stampaj_stampace_btn = QPushButton("🖨️ " + T.get("btn_preview_print", self.lang))
        self.stampaj_stampace_btn.setStyleSheet("background-color: #2196F3; color: white; font-weight: bold;")
        
        self.add_stampac_btn.clicked.connect(self.add_stampac)
        self.edit_stampac_btn.clicked.connect(self.edit_stampac)
        self.delete_stampac_btn.clicked.connect(self.delete_stampac)
        self.excel_stampaci_btn.clicked.connect(self.export_stampace_excel)
        self.stampaj_stampace_btn.clicked.connect(self.stampaj_stampace)
        
        stampaci_btn_layout.addWidget(self.add_stampac_btn)
        stampaci_btn_layout.addWidget(self.edit_stampac_btn)
        stampaci_btn_layout.addWidget(self.delete_stampac_btn)
        stampaci_btn_layout.addWidget(QLabel(T.get("filter_label", self.lang)))
        stampaci_btn_layout.addWidget(self.status_filter)
        stampaci_btn_layout.addWidget(self.ukupno_stampaca_label)  # Dodaj label
        stampaci_btn_layout.addStretch()
        stampaci_btn_layout.addWidget(self.excel_stampaci_btn)
        stampaci_btn_layout.addWidget(self.stampaj_stampace_btn)
        
        self.stampaci_table = QTableWidget()
        self.stampaci_table.setColumnCount(8)
        self.stampaci_table.setHorizontalHeaderLabels([
            T.get("col_id", self.lang), 
            T.get("col_model", self.lang), 
            T.get("col_quantity", self.lang),
            T.get("col_assigned", self.lang),
            T.get("col_available", self.lang),
            T.get("col_status", self.lang), 
            T.get("col_notes", self.lang),
            T.get("col_driver_link", self.lang)
        ])
        self.stampaci_table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.stampaci_table.horizontalHeader().setStretchLastSection(True)
        self.stampaci_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.stampaci_table.cellClicked.connect(self.on_stampac_cell_clicked)
        self.stampaci_table.itemChanged.connect(self.on_stampac_item_changed)
        
        stampaci_layout.addLayout(stampaci_btn_layout)
        stampaci_layout.addWidget(self.stampaci_table)
        
        # Tab 3: RADNICI
        self.radnici_tab = QWidget()
        radnici_layout = QVBoxLayout(self.radnici_tab)
        
        radnici_btn_layout = QHBoxLayout()
        self.add_radnik_btn = QPushButton(T.get("btn_add_employee", self.lang))
        self.edit_radnik_btn = QPushButton(T.get("btn_edit_employee", self.lang))
        self.delete_radnik_btn = QPushButton(T.get("btn_delete_employee", self.lang))
        
        self.add_radnik_btn.clicked.connect(self.add_radnik)
        self.edit_radnik_btn.clicked.connect(self.edit_radnik)
        self.delete_radnik_btn.clicked.connect(self.delete_radnik)
        
        radnici_btn_layout.addWidget(self.add_radnik_btn)
        radnici_btn_layout.addWidget(self.edit_radnik_btn)
        radnici_btn_layout.addWidget(self.delete_radnik_btn)
        radnici_btn_layout.addStretch()
        
        self.radnici_table = QTableWidget()
        self.radnici_table.setColumnCount(3)
        self.radnici_table.setHorizontalHeaderLabels([T.get("col_id", self.lang), T.get("col_first_name", self.lang), T.get("col_last_name", self.lang)])
        self.radnici_table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.radnici_table.horizontalHeader().setStretchLastSection(True)
        self.radnici_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.radnici_table.itemChanged.connect(self.on_radnik_item_changed)
        
        radnici_layout.addLayout(radnici_btn_layout)
        radnici_layout.addWidget(self.radnici_table)
        
        # Tab 4: PREGLED (ko ima šta)
        self.pregled_tab = QWidget()
        pregled_layout = QVBoxLayout(self.pregled_tab)
        
        pregled_info = QLabel(T.get("overview_info", self.lang))
        pregled_info.setStyleSheet("font-style: italic; padding: 5px;")
        pregled_layout.addWidget(pregled_info)
        
        # Dugme za prikaz i štampanje
        pregled_btn_layout = QHBoxLayout()
        preview_pregled_btn = QPushButton(T.get("btn_preview_print", self.lang))
        preview_pregled_btn.clicked.connect(self.preview_pregled)
        preview_pregled_btn.setMaximumWidth(250)
        excel_pregled_btn = QPushButton(T.get("btn_excel_export", self.lang))
        excel_pregled_btn.clicked.connect(self.export_pregled_excel)
        excel_pregled_btn.setMaximumWidth(250)
        pregled_btn_layout.addWidget(preview_pregled_btn)
        pregled_btn_layout.addWidget(excel_pregled_btn)
        pregled_btn_layout.addStretch()
        pregled_layout.addLayout(pregled_btn_layout)
        
        self.pregled_table = QTableWidget()
        self.pregled_table.setColumnCount(4)
        self.pregled_table.setHorizontalHeaderLabels([T.get("col_employee", self.lang), T.get("col_printer", self.lang), T.get("col_status", self.lang), T.get("col_toners", self.lang)])
        self.pregled_table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.pregled_table.horizontalHeader().setStretchLastSection(True)
        self.pregled_table.setSelectionBehavior(QTableWidget.SelectRows)
        
        pregled_layout.addWidget(self.pregled_table)
        
        # Dodaj tabove
        self.tabs.addTab(self.toneri_tab, T.get('tab_toneri', self.lang))
        self.tabs.addTab(self.stampaci_tab, T.get('tab_stampaci', self.lang))
        self.tabs.addTab(self.radnici_tab, T.get('tab_radnici', self.lang))
        self.tabs.addTab(self.pregled_tab, T.get('tab_pregled', self.lang))
        
        # Tab 5: STATISTIKA
        self.statistika_tab = self.create_statistika_tab()
        self.tabs.addTab(self.statistika_tab, T.get('tab_statistika', self.lang))
        
        # Tab 6: ISTORIJA
        self.istorija_tab = self.create_istorija_tab()
        self.tabs.addTab(self.istorija_tab, T.get('tab_istorija', self.lang))
        
        # Tab 7: BACKUP
        self.backup_tab = self.create_backup_tab()
        self.tabs.addTab(self.backup_tab, T.get('tab_backup', self.lang))
        
        main_layout.addWidget(self.tabs)
    
    def create_statistika_tab(self):
        """Kreira tab sa statistikom"""
        tab = QWidget()
        layout = QVBoxLayout(tab)
        
        title = QLabel(T.get('stats_title', self.lang))
        title.setStyleSheet("font-size: 14pt; font-weight: bold; padding: 10px;")
        layout.addWidget(title)
        
        # Period selection
        period_layout = QHBoxLayout()
        period_label = QLabel(T.get('stats_select_period', self.lang))
        period_label.setStyleSheet("font-weight: bold; font-size: 11pt;")
        period_layout.addWidget(period_label)
        
        # Year dropdown
        year_label = QLabel(T.get('stats_year', self.lang))
        self.stats_year_combo = QComboBox()
        from datetime import datetime
        current_year = datetime.now().year
        years = [T.get('stats_all_time', self.lang)] + [str(y) for y in range(2020, current_year + 2)]
        self.stats_year_combo.addItems(years)
        self.stats_year_combo.setCurrentText(str(current_year))
        period_layout.addWidget(year_label)
        period_layout.addWidget(self.stats_year_combo)
        
        # Month dropdown
        month_label = QLabel(T.get('stats_month', self.lang))
        self.stats_month_combo = QComboBox()
        months = [T.get('stats_all_time', self.lang)]
        for i in range(1, 13):
            months.append(T.get(f'month_{i}', self.lang))
        self.stats_month_combo.addItems(months)
        period_layout.addWidget(month_label)
        period_layout.addWidget(self.stats_month_combo)
        
        period_layout.addStretch()
        layout.addLayout(period_layout)
        
        # Dugme za refresh statistike
        refresh_btn = QPushButton(T.get('btn_refresh_stats', self.lang))
        refresh_btn.clicked.connect(self.show_statistika)
        refresh_btn.setMaximumWidth(200)
        layout.addWidget(refresh_btn)
        
        # Placeholder za statistiku
        self.statistika_text = QLabel(T.get('stats_click_refresh', self.lang))
        self.statistika_text.setStyleSheet("padding: 20px; font-size: 11pt;")
        layout.addWidget(self.statistika_text)
        
        layout.addStretch()
        return tab
    
    def create_istorija_tab(self):
        """Kreira tab sa istorijom narudžbina"""
        tab = QWidget()
        layout = QVBoxLayout(tab)
        
        title = QLabel(T.get("history_title", self.lang))
        title.setStyleSheet("font-size: 14pt; font-weight: bold; padding: 10px;")
        layout.addWidget(title)
        
        # Filter za period
        filter_layout = QHBoxLayout()
        filter_layout.addWidget(QLabel(T.get("stats_select_period", self.lang)))
        
        # Godina dropdown
        filter_layout.addWidget(QLabel(T.get("stats_year", self.lang)))
        self.history_year_combo = QComboBox()
        from datetime import datetime
        current_year = datetime.now().year
        years = [T.get("stats_all_time", self.lang)] + [str(y) for y in range(current_year, current_year - 3, -1)]
        self.history_year_combo.addItems(years)
        self.history_year_combo.setCurrentIndex(1)  # Default: current year
        filter_layout.addWidget(self.history_year_combo)
        
        # Mesec dropdown
        filter_layout.addWidget(QLabel(T.get("stats_month", self.lang)))
        self.history_month_combo = QComboBox()
        months = [T.get("stats_all_time", self.lang)] + [T.get(f"month_{i}", self.lang) for i in range(1, 13)]
        self.history_month_combo.addItems(months)
        self.history_month_combo.setCurrentIndex(datetime.now().month)  # Default: current month
        filter_layout.addWidget(self.history_month_combo)
        
        filter_layout.addStretch()
        layout.addLayout(filter_layout)
        
        # Dugmad
        btn_layout = QHBoxLayout()
        add_btn = QPushButton(T.get("btn_add_order", self.lang))
        add_btn.clicked.connect(self.dodaj_narudzbu_u_istoriju)
        refresh_btn = QPushButton(T.get("btn_refresh", self.lang))
        refresh_btn.clicked.connect(self.load_istorija)
        btn_layout.addWidget(add_btn)
        btn_layout.addWidget(refresh_btn)
        btn_layout.addStretch()
        layout.addLayout(btn_layout)
        
        # Tabela
        self.istorija_table = QTableWidget()
        self.istorija_table.setColumnCount(5)
        self.istorija_table.setHorizontalHeaderLabels([T.get("col_id", self.lang), T.get("col_date", self.lang), T.get("col_toner", self.lang), T.get("col_quantity", self.lang), T.get("col_notes", self.lang)])
        self.istorija_table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.istorija_table.horizontalHeader().setStretchLastSection(True)
        layout.addWidget(self.istorija_table)
        
        return tab
    
    def create_backup_tab(self):
        """Kreira tab sa backup opcijama"""
        tab = QWidget()
        layout = QVBoxLayout(tab)
        
        title = QLabel(T.get("backup_title", self.lang))
        title.setStyleSheet("font-size: 14pt; font-weight: bold; padding: 10px;")
        layout.addWidget(title)
        
        # Manuelni backup
        manual_group = QWidget()
        manual_layout = QVBoxLayout(manual_group)
        manual_label = QLabel(T.get("backup_manual", self.lang))
        manual_label.setStyleSheet("font-weight: bold; font-size: 12pt;")
        manual_layout.addWidget(manual_label)
        
        backup_btn = QPushButton(T.get("btn_create_backup", self.lang))
        backup_btn.clicked.connect(self.create_backup)
        backup_btn.setMaximumWidth(250)
        manual_layout.addWidget(backup_btn)
        
        restore_btn = QPushButton(T.get("btn_restore_backup", self.lang))
        restore_btn.clicked.connect(self.restore_backup)
        restore_btn.setMaximumWidth(250)
        manual_layout.addWidget(restore_btn)
        
        layout.addWidget(manual_group)
        
        # Automatski backup
        auto_group = QWidget()
        auto_layout = QVBoxLayout(auto_group)
        auto_label = QLabel(T.get("backup_auto", self.lang))
        auto_label.setStyleSheet("font-weight: bold; font-size: 12pt; margin-top: 20px;")
        auto_layout.addWidget(auto_label)
        
        # Checkbox za enable/disable
        self.auto_backup_checkbox = QCheckBox(T.get("backup_enable", self.lang))
        self.auto_backup_checkbox.stateChanged.connect(self.toggle_auto_backup)
        auto_layout.addWidget(self.auto_backup_checkbox)
        
        # Day of month selector
        day_layout = QHBoxLayout()
        day_label = QLabel(T.get("backup_day", self.lang))
        self.backup_day_spin = QSpinBox()
        self.backup_day_spin.setMinimum(1)
        self.backup_day_spin.setMaximum(28)
        self.backup_day_spin.setValue(1)
        self.backup_day_spin.valueChanged.connect(self.save_backup_settings)
        day_layout.addWidget(day_label)
        day_layout.addWidget(self.backup_day_spin)
        day_layout.addStretch()
        auto_layout.addLayout(day_layout)
        
        # Info o poslednjem backup-u
        self.last_backup_label = QLabel(T.get("backup_last", self.lang) + " " + T.get("backup_never", self.lang))
        self.last_backup_label.setStyleSheet("font-style: italic; padding: 10px;")
        auto_layout.addWidget(self.last_backup_label)
        
        layout.addWidget(auto_group)
        
        # Učitaj trenutne settings
        self.load_backup_settings()
        
        layout.addStretch()
        return tab
    
    def show_statistika(self):
        """Prikazuje statistiku potrošnje za izabrani period"""
        try:
            conn = self.db.get_connection()
            cursor = conn.cursor()
            
            # Get selected period
            selected_year = self.stats_year_combo.currentText()
            selected_month_idx = self.stats_month_combo.currentIndex()
            
            # Build date filter
            date_filter = ""
            period_text = T.get('stats_all_time', self.lang)
            
            if selected_year != T.get('stats_all_time', self.lang):
                if selected_month_idx == 0:  # Whole year
                    date_filter = f"AND strftime('%Y', datum) = '{selected_year}'"
                    period_text = selected_year
                else:  # Specific month
                    month_num = str(selected_month_idx).zfill(2)
                    date_filter = f"AND strftime('%Y-%m', datum) = '{selected_year}-{month_num}'"
                    period_text = f"{T.get(f'month_{selected_month_idx}', self.lang)} {selected_year}"
            
            # Ukupan broj tonera
            cursor.execute("SELECT COUNT(*) FROM toneri")
            ukupno_tonera = cursor.fetchone()[0]
            
            # Toneri ispod minimuma
            cursor.execute("SELECT COUNT(*) FROM toneri WHERE trenutno_stanje < minimalna_kolicina")
            ispod_minimuma = cursor.fetchone()[0]
            
            # Ukupna vrednost stanja
            cursor.execute("SELECT SUM(trenutno_stanje) FROM toneri")
            ukupno_stanje = cursor.fetchone()[0] or 0
            
            # Top 5 tonera po stanju
            cursor.execute("""
                SELECT model, trenutno_stanje 
                FROM toneri 
                ORDER BY trenutno_stanje DESC 
                LIMIT 5
            """)
            top_toneri = cursor.fetchall()
            
            # Potrošnja za izabrani period
            query = f"SELECT COUNT(*) FROM istorija_potrosnje WHERE 1=1 {date_filter}"
            cursor.execute(query)
            potrosnja_period = cursor.fetchone()[0]
            
            conn.close()
            
            # Formatiraj prikaz
            stats_text = f"""
<h2>{T.get('stats_general', self.lang)}</h2>
<ul>
    <li><b>{T.get('stats_total_toners', self.lang)}</b> {ukupno_tonera}</li>
    <li><b>{T.get('stats_below_min', self.lang)}</b> <span style="color: red;">{ispod_minimuma}</span></li>
    <li><b>{T.get('stats_total_stock', self.lang)}</b> {ukupno_stanje}</li>
    <li><b>{T.get('stats_consumption', self.lang)}</b> {potrosnja_period} {T.get('stats_pcs', self.lang)} ({period_text})</li>
</ul>

<h2>{T.get('stats_top5', self.lang)}</h2>
<ol>
"""
            for model, stanje in top_toneri:
                stats_text += f"    <li><b>{model}</b>: {stanje} {T.get('stats_pcs', self.lang)}</li>\n"
            
            stats_text += "</ol>"
            
            self.statistika_text.setText(stats_text)
            self.statistika_text.setWordWrap(True)
            
        except Exception as e:
            QMessageBox.critical(self, T.get('error', self.lang), f"{T.get('error', self.lang)}:\n{str(e)}")
    
    def load_istorija(self):
        """Učitava istorija narudžbina sa filterom po periodu"""
        conn = self.db.get_connection()
        cursor = conn.cursor()
        
        # Get selected period
        selected_year = self.history_year_combo.currentText()
        selected_month = self.history_month_combo.currentIndex()
        
        # Build query with date filter
        query = """
            SELECT 
                i.id,
                i.datum,
                t.model,
                i.kolicina,
                i.napomena
            FROM istorija_narudzbi i
            LEFT JOIN toneri t ON i.toner_id = t.id
        """
        
        params = []
        where_clauses = []
        
        # Year filter
        if selected_year != T.get("stats_all_time", self.lang):
            where_clauses.append("strftime('%Y', i.datum) = ?")
            params.append(selected_year)
        
        # Month filter
        if selected_month > 0:  # 0 is "All Time"
            where_clauses.append("strftime('%m', i.datum) = ?")
            params.append(f"{selected_month:02d}")
        
        if where_clauses:
            query += " WHERE " + " AND ".join(where_clauses)
        
        query += " ORDER BY i.datum DESC"
        
        cursor.execute(query, params)
        rows = cursor.fetchall()
        conn.close()
        
        self.istorija_table.setRowCount(len(rows))
        for i, row in enumerate(rows):
            for j, value in enumerate(row):
                # Translate napomena column (j==4)
                if j == 4 and value:
                    note_map = {
                        'Automatska narudžbina': T.get('auto_order_note', self.lang),
                        'Narudžbina': T.get('manual_order_note', self.lang)
                    }
                    display_value = note_map.get(value, value)
                    item = QTableWidgetItem(display_value)
                else:
                    item = QTableWidgetItem(str(value) if value else "")
                self.istorija_table.setItem(i, j, item)
    
    def dodaj_narudzbu_u_istoriju(self):
        """Dodaje trenutnu narudžbinu u istoriju"""
        try:
            conn = self.db.get_connection()
            cursor = conn.cursor()
            
            # Pronađi sve tonere ispod minimuma
            cursor.execute("""
                SELECT id, model, (minimalna_kolicina - trenutno_stanje) as kolicina
                FROM toneri 
                WHERE CAST(trenutno_stanje AS INTEGER) < CAST(minimalna_kolicina AS INTEGER)
            """)
            
            rows = cursor.fetchall()
            
            if not rows:
                QMessageBox.information(self, T.get("info", self.lang), T.get("info_no_orders", self.lang))
                return
            
            # Dodaj svaki toner u istoriju
            from datetime import datetime
            datum = datetime.now().strftime('%Y-%m-%d')
            
            for toner_id, model, kolicina in rows:
                cursor.execute("""
                    INSERT INTO istorija_narudzbi (datum, toner_id, kolicina, napomena)
                    VALUES (?, ?, ?, ?)
                """, (datum, toner_id, kolicina, T.get("auto_order_note", self.lang) if hasattr(self, "lang") else "Automatska narudžbina"))
            
            conn.commit()
            conn.close()
            
            self.load_istorija()
            QMessageBox.information(self, T.get("success", self.lang), f"Dodato {len(rows)} stavki u istoriju narudžbina!")
            
        except Exception as e:
            QMessageBox.critical(self, T.get("error", self.lang), f"Greška:\n{str(e)}")
    
    def dodaj_narudzbu_dialog(self, dialog, rows):
        """Dodaje narudžbinu u istoriju iz dialoga i zatvara dialog"""
        try:
            conn = self.db.get_connection()
            cursor = conn.cursor()
            
            from datetime import datetime
            datum = datetime.now().strftime('%Y-%m-%d')
            
            # Konvertuj rows u format (toner_id, model, kolicina)
            for row in rows:
                model, min_kol, trenutno, za_nar = row
                # Nađi toner_id po modelu
                cursor.execute("SELECT id FROM toneri WHERE model = ?", (model,))
                result = cursor.fetchone()
                if result:
                    toner_id = result[0]
                    cursor.execute("""
                        INSERT INTO istorija_narudzbi (datum, toner_id, kolicina, napomena)
                        VALUES (?, ?, ?, ?)
                    """, (datum, toner_id, za_nar, T.get("manual_order_note", self.lang) if hasattr(self, "lang") else "Narudžbina"))
            
            conn.commit()
            conn.close()
            
            self.load_istorija()
            QMessageBox.information(self, T.get("success", self.lang), T.get("msg_added_to_history", self.lang).format(len(rows)))
            dialog.close()
            
        except Exception as e:
            QMessageBox.critical(self, T.get("error", self.lang), f"Greška:\n{str(e)}")
    
    def load_backup_settings(self):
        """Učitava backup settings iz baze"""
        conn = self.db.get_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT enabled, day_of_month, last_backup_date FROM backup_settings WHERE id = 1")
        row = cursor.fetchone()
        conn.close()
        
        if row:
            enabled, day, last_date = row
            self.auto_backup_checkbox.setChecked(bool(enabled))
            self.backup_day_spin.setValue(day or 1)
            if last_date:
                self.last_backup_label.setText(f"Poslednji backup: {last_date}")
    
    def toggle_auto_backup(self, state):
        """Uključuje/isključuje automatski backup"""
        self.save_backup_settings()
    
    def save_backup_settings(self):
        """Čuva backup settings"""
        conn = self.db.get_connection()
        cursor = conn.cursor()
        cursor.execute("""
            UPDATE backup_settings 
            SET enabled = ?, day_of_month = ?
            WHERE id = 1
        """, (1 if self.auto_backup_checkbox.isChecked() else 0, self.backup_day_spin.value()))
        conn.commit()
        conn.close()
    
    def cleanup_old_history(self):
        """Automatski briše zapise iz istorije starije od 2 godine"""
        try:
            conn = self.db.get_connection()
            cursor = conn.cursor()
            
            # Obriši zapise starije od 2 godine
            cursor.execute("""
                DELETE FROM istorija_narudzbi 
                WHERE datum < date('now', '-2 years')
            """)
            
            deleted_count = cursor.rowcount
            conn.commit()
            conn.close()
            
            if deleted_count > 0:
                print(f"🧹 Cleaned up {deleted_count} old history records (older than 2 years)")
        except Exception as e:
            print(f"Cleanup error: {e}")
    
    def check_auto_backup(self):
        """Proverava da li treba izvršiti automatski backup"""
        try:
            conn = self.db.get_connection()
            cursor = conn.cursor()
            cursor.execute("SELECT enabled, day_of_month, last_backup_date FROM backup_settings WHERE id = 1")
            row = cursor.fetchone()
            conn.close()
            
            if not row or not row[0]:
                return  # Nije omogućeno
            
            enabled, target_day, last_backup = row
            
            from datetime import datetime
            today = datetime.now()
            
            # Proveri da li je prošao ciljni dan ovog meseca
            if today.day >= target_day:
                # Proveri da li je već backup urađen ovog meseca
                if last_backup:
                    last_backup_date = datetime.strptime(last_backup, '%Y-%m-%d')
                    if last_backup_date.month == today.month and last_backup_date.year == today.year:
                        return  # Već urađen ovog meseca
                
                # Uradi backup (prvi put ovog meseca posle ciljnog dana)
                self.create_backup(auto=True)
        except Exception as e:
            print(f"Auto backup error: {e}")
    
    def create_backup(self, auto=False):
        """Kreira backup baze podataka"""
        try:
            import shutil
            from datetime import datetime
            import os
            
            # Zatvori konekciju pre backup-a
            db_path = "toneri.db"
            
            if not os.path.exists(db_path):
                QMessageBox.warning(self, T.get("error", self.lang), T.get("error_db_not_found", self.lang))
                return
            
            # Kreiraj backup folder ako ne postoji
            backup_folder = "backups"
            os.makedirs(backup_folder, exist_ok=True)
            
            # Ime backup fajla
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            backup_name = f"toneri_backup_{timestamp}.db"
            backup_path = os.path.join(backup_folder, backup_name)
            
            # Kopiraj bazu
            shutil.copy2(db_path, backup_path)
            
            # Ažuriraj last_backup_date
            conn = self.db.get_connection()
            cursor = conn.cursor()
            cursor.execute("UPDATE backup_settings SET last_backup_date = date('now') WHERE id = 1")
            conn.commit()
            conn.close()
            
            self.load_backup_settings()
            
            if not auto:
                QMessageBox.information(self, T.get("success", self.lang), 
                    T.get("msg_backup_created", self.lang) + "\n\n" + 
                    T.get("msg_file", self.lang) + " " + backup_name + "\n" + 
                    T.get("msg_location", self.lang) + " " + os.path.abspath(backup_path))
            
        except Exception as e:
            QMessageBox.critical(self, T.get("error", self.lang), f"Greška prilikom backup-a:\n{str(e)}")
    
    def restore_backup(self):
        """Restoruje bazu iz backup-a"""
        try:
            import os
            import shutil
            from datetime import datetime
            from PyQt5.QtWidgets import QFileDialog
            
            # Otvori file dialog
            file_path, _ = QFileDialog.getOpenFileName(
                self, 
                T.get("backup_select_file", self.lang), 
                "backups", 
                "Database files (*.db)"
            )
            
            if not file_path:
                return
            
            reply = QMessageBox.question(
                self, 
                T.get("confirm", self.lang), 
                T.get("confirm_restore", self.lang),
                QMessageBox.Yes | QMessageBox.No
            )
            
            if reply == QMessageBox.Yes:
                # Zatvori sve konekcije
                db_path = "toneri.db"
                
                # Backup trenutne baze pre restore-a
                backup_current = f"toneri_pre_restore_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db"
                shutil.copy2(db_path, os.path.join("backups", backup_current))
                
                # Restore
                shutil.copy2(file_path, db_path)
                
                QMessageBox.information(self, T.get("success", self.lang), T.get("msg_db_restored", self.lang))
                
                # Close application (user needs to restart manually)
                # Note: Auto-restart doesn't work well with PyInstaller EXE
                self.close()
                
        except Exception as e:
            QMessageBox.critical(self, T.get("error", self.lang), f"Greška prilikom restore-a:\n{str(e)}")

    def load_all_data(self):
        # Privremeno blokiraj search signal da ne bi došlo do rekurzije
        try:
            self.search_input.blockSignals(True)
            self.load_toneri()
            self.load_stampaci()
            self.load_radnici()
            self.load_pregled()
            self.load_istorija()
            # Re-apply search if active
            if hasattr(self, 'search_input') and self.search_input.text():
                self.search_all()
        finally:
            self.search_input.blockSignals(False)
    
    def save_column_widths(self):
        """Čuva širine kolona svih tabela"""
        try:
            # Toneri
            for i in range(self.toneri_table.columnCount()):
                self.settings.setValue(f'toneri_col_{i}_width', self.toneri_table.columnWidth(i))
            
            # Štampači
            for i in range(self.stampaci_table.columnCount()):
                self.settings.setValue(f'stampaci_col_{i}_width', self.stampaci_table.columnWidth(i))
            
            # Radnici
            for i in range(self.radnici_table.columnCount()):
                self.settings.setValue(f'radnici_col_{i}_width', self.radnici_table.columnWidth(i))
            
            # Pregled
            for i in range(self.pregled_table.columnCount()):
                self.settings.setValue(f'pregled_col_{i}_width', self.pregled_table.columnWidth(i))
            
            # Istorija
            if hasattr(self, 'istorija_table'):
                for i in range(self.istorija_table.columnCount()):
                    self.settings.setValue(f'istorija_col_{i}_width', self.istorija_table.columnWidth(i))
        except Exception as e:
            print(f"Error saving column widths: {e}")
    
    def restore_column_widths(self):
        """Učitava sačuvane širine kolona"""
        try:
            # Toneri
            for i in range(self.toneri_table.columnCount()):
                width = self.settings.value(f'toneri_col_{i}_width', None)
                if width:
                    self.toneri_table.setColumnWidth(i, int(width))
            
            # Štampači
            for i in range(self.stampaci_table.columnCount()):
                width = self.settings.value(f'stampaci_col_{i}_width', None)
                if width:
                    self.stampaci_table.setColumnWidth(i, int(width))
            
            # Radnici
            for i in range(self.radnici_table.columnCount()):
                width = self.settings.value(f'radnici_col_{i}_width', None)
                if width:
                    self.radnici_table.setColumnWidth(i, int(width))
            
            # Pregled
            for i in range(self.pregled_table.columnCount()):
                width = self.settings.value(f'pregled_col_{i}_width', None)
                if width:
                    self.pregled_table.setColumnWidth(i, int(width))
            
            # Istorija
            if hasattr(self, 'istorija_table'):
                for i in range(self.istorija_table.columnCount()):
                    width = self.settings.value(f'istorija_col_{i}_width', None)
                    if width:
                        self.istorija_table.setColumnWidth(i, int(width))
        except Exception as e:
            print(f"Error restoring column widths: {e}")
    
    def closeEvent(self, event):
        """Override closeEvent da sačuva širine kolona pri zatvaranju"""
        self.save_column_widths()
        event.accept()
    
    def update_ukupno_tonera(self):
        """Računa i prikazuje ukupan zbir svih tonera u realnom vremenu"""
        try:
            conn = self.db.get_connection()
            cursor = conn.cursor()
            cursor.execute("SELECT COALESCE(SUM(trenutno_stanje), 0) FROM toneri")
            ukupno = cursor.fetchone()[0]
            conn.close()
            
            if self.lang == 'sr':
                self.ukupno_tonera_label.setText(f"📦 Ukupno tonera: {ukupno}")
            else:
                self.ukupno_tonera_label.setText(f"📦 Total toners: {ukupno}")
        except Exception as e:
            print(f"Error updating total toners: {e}")
            self.ukupno_tonera_label.setText("📦 Ukupno: 0")
    
    def update_ukupno_stampaca(self):
        """Računa i prikazuje ukupan broj svih štampača u realnom vremenu"""
        try:
            conn = self.db.get_connection()
            cursor = conn.cursor()
            cursor.execute("SELECT COALESCE(SUM(kolicina), 0) FROM stampaci")
            ukupno = cursor.fetchone()[0]
            conn.close()
            
            if self.lang == 'sr':
                self.ukupno_stampaca_label.setText(f"🖨️ Ukupno štampača: {ukupno}")
            else:
                self.ukupno_stampaca_label.setText(f"🖨️ Total printers: {ukupno}")
        except Exception as e:
            print(f"Error updating total printers: {e}")
            self.ukupno_stampaca_label.setText("🖨️ Ukupno: 0")
    
    def load_pregled(self):
        """Učitava kompletan pregled: SVE radnike, štampače i tonere - označava nepovezane crveno"""
        conn = self.db.get_connection()
        cursor = conn.cursor()
        
        all_rows = []
        
        # 1. Sve veze radnik-stampac-toneri (kompletno povezani)
        cursor.execute('''
            SELECT 
                r.ime || ' ' || r.prezime as radnik,
                s.model as stampac_model,
                s.status,
                GROUP_CONCAT(t.model, ', ') as toneri
            FROM radnici r
            JOIN radnik_stampaci rs ON r.id = rs.radnik_id
            JOIN stampaci s ON rs.stampac_id = s.id
            LEFT JOIN stampac_toneri st ON s.id = st.stampac_id
            LEFT JOIN toneri t ON st.toner_id = t.id
            GROUP BY r.id, s.id
            ORDER BY r.prezime, r.ime, s.model
        ''')
        all_rows.extend(cursor.fetchall())
        
        # 2. Radnici BEZ štampača
        cursor.execute('''
            SELECT 
                r.ime || ' ' || r.prezime as radnik,
                NULL as stampac_model,
                NULL as status,
                NULL as toneri
            FROM radnici r
            WHERE r.id NOT IN (SELECT DISTINCT radnik_id FROM radnik_stampaci)
            ORDER BY r.prezime, r.ime
        ''')
        all_rows.extend(cursor.fetchall())
        
        # 3. Štampači BEZ radnika (ali možda sa tonerima)
        cursor.execute('''
            SELECT 
                NULL as radnik,
                s.model as stampac_model,
                s.status,
                GROUP_CONCAT(t.model, ', ') as toneri
            FROM stampaci s
            LEFT JOIN stampac_toneri st ON s.id = st.stampac_id
            LEFT JOIN toneri t ON st.toner_id = t.id
            WHERE s.id NOT IN (SELECT DISTINCT stampac_id FROM radnik_stampaci WHERE stampac_id IS NOT NULL)
            GROUP BY s.id
            ORDER BY s.model
        ''')
        all_rows.extend(cursor.fetchall())
        
        # 4. Toneri BEZ štampača
        cursor.execute('''
            SELECT 
                NULL as radnik,
                NULL as stampac_model,
                NULL as status,
                t.model as toneri
            FROM toneri t
            WHERE t.id NOT IN (SELECT DISTINCT toner_id FROM stampac_toneri WHERE toner_id IS NOT NULL)
            ORDER BY t.model
        ''')
        all_rows.extend(cursor.fetchall())
        
        conn.close()
        
        self.pregled_table.setRowCount(len(all_rows))
        for i, row in enumerate(all_rows):
            for j, value in enumerate(row):
                # Determine if this field is missing (NULL or None)
                is_missing = value is None or value == "" or value == "None"
                
                # Translate status column (j==2)
                if j == 2 and value and not is_missing:
                    status_map = {
                        'Aktivan': T.get('status_active', self.lang),
                        'Na servisu': T.get('status_in_service', self.lang),
                        'Za rashod': T.get('status_for_disposal', self.lang)
                    }
                    display_value = status_map.get(value, value)
                    item = QTableWidgetItem(display_value)
                else:
                    item = QTableWidgetItem(str(value) if value and not is_missing else "-")
                
                # Color logic
                if is_missing:
                    # MISSING data - RED background
                    item.setBackground(QColor(255, 200, 200))
                    item.setForeground(QColor(0, 0, 0))  # Black text
                elif j == 2:  # Status column with data
                    # Check against database values for coloring
                    if value == "Na servisu":
                        item.setBackground(QColor(255, 255, 200))
                    elif value == "Za rashod":
                        item.setBackground(QColor(255, 150, 150))
                    else:
                        item.setBackground(QColor(255, 255, 255))  # White for normal status
                    item.setForeground(QColor(0, 0, 0))  # Black text
                else:
                    # Normal data - white background
                    item.setBackground(QColor(255, 255, 255))
                    item.setForeground(QColor(0, 0, 0))  # Black text
                
                # Zaključaj sve ćelije - PREGLED je read-only
                item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                
                self.pregled_table.setItem(i, j, item)
    
        # Re-apply search highlighting
        if hasattr(self, "search_input") and self.search_input.text():
            self.apply_search_highlighting()

    def on_toner_stanje_changed(self, item):
        """Kada korisnik promeni vrednost u bilo kojoj koloni tonera"""
        col = item.column()
        
        if col == 0:  # ID - ne menjaj
            return
        
        row = item.row()
        toner_id = int(self.toneri_table.item(row, 0).text())
        
        conn = self.db.get_connection()
        cursor = conn.cursor()
        
        try:
            if col == 1:  # Model
                new_value = item.text().strip()
                if not new_value:
                    QMessageBox.warning(self, T.get("error", self.lang), "Model ne može biti prazan!")
                    self.load_toneri()
                    return
                cursor.execute("UPDATE toneri SET model = ? WHERE id = ?", (new_value, toner_id))
            
            elif col == 2:  # Min. količina
                new_value = int(item.text())
                if new_value < 0:
                    QMessageBox.warning(self, T.get("error", self.lang), "Minimalna količina ne može biti negativna!")
                    self.load_toneri()
                    return
                cursor.execute("UPDATE toneri SET minimalna_kolicina = ? WHERE id = ?", (new_value, toner_id))
            
            elif col == 3:  # Stanje
                new_value = int(item.text())
                if new_value < 0:
                    QMessageBox.warning(self, T.get("error", self.lang), "Stanje ne može biti negativno!")
                    self.load_toneri()
                    return
                cursor.execute("UPDATE toneri SET trenutno_stanje = ? WHERE id = ?", (new_value, toner_id))
            
            conn.commit()
            conn.close()
            
            # Osvježi prikaz
            try:
                self.toneri_table.itemChanged.disconnect(self.on_toner_stanje_changed)
            except:
                pass
            self.load_toneri()
            
        except ValueError:
            QMessageBox.warning(self, T.get("error", self.lang), T.get("error_must_be_number", self.lang))
            conn.close()
            self.load_toneri()
        except Exception as e:
            QMessageBox.critical(self, T.get("error", self.lang), f"Greška prilikom čuvanja: {str(e)}")
            conn.close()
            self.load_toneri()
    
    def on_stampac_item_changed(self, item):
        """Kada korisnik promeni vrednost u bilo kojoj koloni štampača"""
        col = item.column()
        
        if col == 0:  # ID - ne menjaj
            return
        
        if col in [3, 4]:  # Dodeljeno, Slobodno - calculated, ne menjaj
            return
        
        if col == 5:  # Status - ima poseban dropdown, preskači
            return
        
        row = item.row()
        stampac_id = int(self.stampaci_table.item(row, 0).text())
        
        conn = self.db.get_connection()
        cursor = conn.cursor()
        
        try:
            if col == 1:  # Model
                new_value = item.text().strip()
                if not new_value:
                    QMessageBox.warning(self, T.get("error", self.lang), "Model ne može biti prazan!")
                    self.load_stampaci()
                    return
                cursor.execute("UPDATE stampaci SET model = ? WHERE id = ?", (new_value, stampac_id))
            
            elif col == 2:  # Količina
                try:
                    new_value = int(item.text().strip())
                    if new_value < 0:
                        raise ValueError("Količina mora biti pozitivna")
                    
                    # Check current quantity
                    cursor.execute("SELECT COALESCE(kolicina, 1) FROM stampaci WHERE id = ?", (stampac_id,))
                    old_kolicina = cursor.fetchone()[0]
                    
                    # Check if reducing below assigned count
                    cursor.execute("""
                        SELECT COUNT(DISTINCT radnik_id) 
                        FROM radnik_stampaci 
                        WHERE stampac_id = ?
                    """, (stampac_id,))
                    dodeljeno = cursor.fetchone()[0]
                    
                    # If reducing quantity AND there are assigned employees, ask what to do
                    if new_value < old_kolicina and dodeljeno > 0:
                        if new_value < dodeljeno:
                            # MUST unassign (new quantity below assigned count)
                            mandatory = True
                            unassign_count = dodeljeno - new_value
                            message = (T.get("reducing_quantity", self.lang).format(old_kolicina, new_value) + "\n\n" +
                                     T.get("must_unassign", self.lang).format(unassign_count) + "\n" +
                                     T.get("choose_employees", self.lang))
                        else:
                            # CAN optionally unassign (new quantity >= assigned count)
                            mandatory = False
                            # Ask if user wants to unassign anyone
                            cursor.execute("""
                                SELECT r.id, r.ime, r.prezime
                                FROM radnici r
                                JOIN radnik_stampaci rs ON r.id = rs.radnik_id
                                WHERE rs.stampac_id = ?
                            """, (stampac_id,))
                            employees = cursor.fetchall()
                            
                            emp_list = "\n".join([f"• {ime} {prezime}" for _, ime, prezime in employees])
                            
                            reply = QMessageBox.question(
                                self,
                                T.get("unassign_question_title", self.lang),
                                T.get("reducing_quantity", self.lang).format(old_kolicina, new_value) + "\n\n" +
                                T.get("currently_assigned", self.lang).format(dodeljeno) + "\n" + emp_list + "\n\n" +
                                T.get("want_to_unassign", self.lang),
                                QMessageBox.Yes | QMessageBox.No
                            )
                            
                            if reply == QMessageBox.No:
                                # Just reduce quantity, keep all assignments
                                cursor.execute("UPDATE stampaci SET kolicina = ? WHERE id = ?", (new_value, stampac_id))
                                conn.commit()
                                conn.close()
                                self.load_stampaci()
                                self.load_pregled()
                                return
                            
                            # User wants to unassign - continue to selection dialog
                            unassign_count = None  # User can choose how many
                            message = (T.get("reducing_quantity", self.lang).format(old_kolicina, new_value) + "\n\n" +
                                     T.get("choose_to_unassign", self.lang).format(dodeljeno))
                        
                        # Get all assigned employees
                        cursor.execute("""
                            SELECT r.id, r.ime, r.prezime
                            FROM radnici r
                            JOIN radnik_stampaci rs ON r.id = rs.radnik_id
                            WHERE rs.stampac_id = ?
                        """, (stampac_id,))
                        employees = cursor.fetchall()
                        
                        # Create selection dialog
                        dialog = QDialog(self)
                        dialog.setWindowTitle(T.get("unassign_dialog_title", self.lang))
                        dialog.setMinimumWidth(400)
                        layout = QVBoxLayout(dialog)
                        
                        info_label = QLabel(message)
                        info_label.setWordWrap(True)
                        layout.addWidget(info_label)
                        
                        # Checkboxes for each employee
                        checkboxes = []
                        for emp_id, ime, prezime in employees:
                            cb = QCheckBox(f"{ime} {prezime}")
                            cb.setProperty("employee_id", emp_id)
                            checkboxes.append(cb)
                            layout.addWidget(cb)
                        
                        # Buttons
                        btn_layout = QHBoxLayout()
                        ok_btn = QPushButton(T.get("btn_unassign", self.lang))
                        cancel_btn = QPushButton(T.get("btn_cancel", self.lang))
                        ok_btn.clicked.connect(dialog.accept)
                        cancel_btn.clicked.connect(dialog.reject)
                        btn_layout.addWidget(ok_btn)
                        btn_layout.addWidget(cancel_btn)
                        layout.addLayout(btn_layout)
                        
                        if dialog.exec_() == QDialog.Accepted:
                            # Get selected employees
                            selected_ids = [cb.property("employee_id") for cb in checkboxes if cb.isChecked()]
                            
                            if mandatory and len(selected_ids) != unassign_count:
                                QMessageBox.warning(self, T.get("error", self.lang), 
                                    T.get("must_select_exactly", self.lang).format(unassign_count))
                                self.load_stampaci()
                                conn.close()
                                return
                            
                            if not mandatory and len(selected_ids) == 0:
                                QMessageBox.warning(self, T.get("error", self.lang), 
                                    T.get("must_select_at_least_one", self.lang))
                                self.load_stampaci()
                                conn.close()
                                return
                            
                            # Unassign selected employees
                            for emp_id in selected_ids:
                                cursor.execute("DELETE FROM radnik_stampaci WHERE radnik_id = ? AND stampac_id = ?", 
                                             (emp_id, stampac_id))
                            
                            # Now update quantity
                            cursor.execute("UPDATE stampaci SET kolicina = ? WHERE id = ?", (new_value, stampac_id))
                        else:
                            self.load_stampaci()
                            conn.close()
                            return
                    else:
                        # No employees assigned OR increasing quantity - just update
                        cursor.execute("UPDATE stampaci SET kolicina = ? WHERE id = ?", (new_value, stampac_id))
                except ValueError:
                    QMessageBox.warning(self, T.get("error", self.lang), "Količina mora biti broj!")
                    self.load_stampaci()
                    return
            
            elif col == 6:  # Napomena
                new_value = item.text().strip()
                cursor.execute("UPDATE stampaci SET napomena = ? WHERE id = ?", (new_value, stampac_id))
            
            elif col == 7:  # Driver Link
                new_value = item.text().strip()
                cursor.execute("UPDATE stampaci SET driver_link = ? WHERE id = ?", (new_value, stampac_id))
            
            conn.commit()
            conn.close()
            
            # Osvježi prikaz
            try:
                self.stampaci_table.itemChanged.disconnect(self.on_stampac_item_changed)
            except:
                pass
            self.load_stampaci()
            self.load_pregled()
            
        except Exception as e:
            QMessageBox.critical(self, T.get("error", self.lang), f"Greška prilikom čuvanja: {str(e)}")
            conn.close()
            self.load_stampaci()
    
    def on_radnik_item_changed(self, item):
        """Kada korisnik promeni vrednost u bilo kojoj koloni radnika"""
        col = item.column()
        
        if col == 0:  # ID - ne menjaj
            return
        
        row = item.row()
        radnik_id = int(self.radnici_table.item(row, 0).text())
        
        conn = self.db.get_connection()
        cursor = conn.cursor()
        
        try:
            if col == 1:  # Ime
                new_value = item.text().strip()
                if not new_value:
                    QMessageBox.warning(self, T.get("error", self.lang), "Ime ne može biti prazno!")
                    self.load_radnici()
                    return
                cursor.execute("UPDATE radnici SET ime = ? WHERE id = ?", (new_value, radnik_id))
            
            elif col == 2:  # Prezime
                new_value = item.text().strip()
                if not new_value:
                    QMessageBox.warning(self, T.get("error", self.lang), "Prezime ne može biti prazno!")
                    self.load_radnici()
                    return
                cursor.execute("UPDATE radnici SET prezime = ? WHERE id = ?", (new_value, radnik_id))
            
            conn.commit()
            conn.close()
            
            # Osvježi prikaz
            try:
                self.radnici_table.itemChanged.disconnect(self.on_radnik_item_changed)
            except:
                pass
            self.load_radnici()
            self.load_pregled()
            
        except Exception as e:
            QMessageBox.critical(self, T.get("error", self.lang), f"Greška prilikom čuvanja: {str(e)}")
            conn.close()
            self.load_radnici()
    
    def on_stampac_cell_clicked(self, row, column):
        """Kada se klikne na kolonu Status (5) ili Driver Link (7)"""
        if column == 5:  # Status kolona (sada na poziciji 5)
            # Prvo proveri da li već postoji combo box u ovoj ćeliji
            existing_widget = self.stampaci_table.cellWidget(row, column)
            if existing_widget:
                # Ako već postoji widget, ukloni ga i ne radi ništa
                self.stampaci_table.removeCellWidget(row, column)
                return
            
            stampac_id = int(self.stampaci_table.item(row, 0).text())
            current_status = self.stampaci_table.item(row, 5).text()
            
            # Kreiraj combo box
            combo = QComboBox()
            combo.addItems([T.get("status_active", self.lang), T.get("status_in_service", self.lang), T.get("status_for_disposal", self.lang)])
            combo.setCurrentText(current_status)
            
            # Postavi combo u ćeliju
            self.stampaci_table.setCellWidget(row, column, combo)
            combo.showPopup()
            
            # Kada se izabere nova vrednost (koristeći activated umesto currentTextChanged)
            def on_status_activated(index):
                new_status = combo.itemText(index)
                
                # Prvo ukloni combo iz tabele
                self.stampaci_table.removeCellWidget(row, column)
                
                if new_status != current_status:
                    reply = QMessageBox.question(self, T.get("confirm", self.lang), 
                                                T.get("confirm_change_status", self.lang).format(new_status),
                                                QMessageBox.Yes | QMessageBox.No)
                    if reply == QMessageBox.Yes:
                        # Konvertuj prevedeni status nazad u srpski za bazu
                        reverse_status_map = {
                            T.get('status_active', self.lang): 'Aktivan',
                            T.get('status_in_service', self.lang): 'Na servisu',
                            T.get('status_for_disposal', self.lang): 'Za rashod'
                        }
                        db_status = reverse_status_map.get(new_status, new_status)
                        
                        # Sačuvaj u bazu
                        conn = self.db.get_connection()
                        cursor = conn.cursor()
                        cursor.execute("UPDATE stampaci SET status = ? WHERE id = ?", 
                                     (db_status, stampac_id))
                        conn.commit()
                        conn.close()
                        
                        # Osvježi tabelu
                        self.load_stampaci()
                        self.load_pregled()
                        QMessageBox.information(self, T.get("success", self.lang), T.get("msg_status_changed", self.lang))
                    else:
                        # Samo reload bez promene
                        self.load_stampaci()
                # Ako je isti status, ne radi ništa - combo je već uklonjen
            
            # Koristi activated signal umesto currentTextChanged
            combo.activated.connect(on_status_activated)
            
            # Dodaj i handler za zatvaranje popup-a bez izbora
            def on_popup_hidden():
                # Ukloni combo ako popup bude zatvoren bez izbora
                if self.stampaci_table.cellWidget(row, column) == combo:
                    self.stampaci_table.removeCellWidget(row, column)
            
            # Ovo nije idealno, ali možemo koristiti timer
            combo.hidePopup_original = combo.hidePopup
            def hidePopup_wrapped():
                combo.hidePopup_original()
                # Odloži uklanjanje da signal 'activated' može da se izvrši prvo
                QTimer.singleShot(100, on_popup_hidden)
            combo.hidePopup = hidePopup_wrapped
        
        elif column == 7:  # Driver Link kolona
            item = self.stampaci_table.item(row, 7)
            if item and item.text().strip():
                link = item.text().strip()
                
                # Ako link ne počinje sa http:// ili https://, dodaj https://
                if not link.startswith('http://') and not link.startswith('https://'):
                    if link.startswith('www.'):
                        link = 'https://' + link
                    else:
                        link = 'https://' + link
                
                # Otvori link pomoću odloženog poziva (sigurnije za Qt)
                def open_link():
                    try:
                        QDesktopServices.openUrl(QUrl(link))
                    except Exception as e:
                        QMessageBox.warning(self, T.get("error", self.lang), f"Ne mogu otvoriti link:\n{str(e)}")
                
                QTimer.singleShot(100, open_link)  # Odloži 100ms
    
    def load_toneri(self):
        conn = self.db.get_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT id, model, minimalna_kolicina, trenutno_stanje FROM toneri ORDER BY model")
        rows = cursor.fetchall()
        conn.close()
        
        # Privremeno onemogući signal
        try:
            self.toneri_table.itemChanged.disconnect(self.on_toner_stanje_changed)
        except:
            pass
        
        self.toneri_table.setRowCount(len(rows))
        for i, row in enumerate(rows):
            for j, value in enumerate(row):
                item = QTableWidgetItem(str(value) if value is not None else "")
                
                # SVE kolone osim ID (0) su editabilne
                if j == 0:
                    item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                else:
                    item.setFlags(item.flags() | Qt.ItemIsEditable)
                
                # Oboji red ako je ispod minimuma - min_kol je j=2, stanje je j=3
                if j == 3 and row[3] < row[2] and not self.search_active:  # stanje < min_količina
                    item.setBackground(QColor(255, 200, 200))
                self.toneri_table.setItem(i, j, item)
        
        # Ponovo omogući signal
        try:
            self.toneri_table.itemChanged.disconnect(self.on_toner_stanje_changed)
        except:
            pass
        self.toneri_table.itemChanged.connect(self.on_toner_stanje_changed)
        
        # Ažuriraj ukupan zbir tonera
        self.update_ukupno_tonera()
        
        # Re-apply search highlighting if search is active
        if hasattr(self, 'search_input') and self.search_input and self.search_input.text():
            self.apply_search_highlighting()
    
    def load_stampaci(self):
        conn = self.db.get_connection()
        cursor = conn.cursor()
        
        status_filter = self.status_filter.currentText()
        
        # Konvertuj prevedeni status nazad u srpski za upit
        reverse_status_map = {
            T.get('status_all', self.lang): 'Svi',
            T.get('status_active', self.lang): 'Aktivan',
            T.get('status_in_service', self.lang): 'Na servisu',
            T.get('status_for_disposal', self.lang): 'Za rashod'
        }
        db_status_filter = reverse_status_map.get(status_filter, status_filter)
        
        # Query with calculated dodeljeno count
        if db_status_filter == 'Svi' or status_filter == T.get("status_all", self.lang):
            cursor.execute("""
                SELECT 
                    s.id, 
                    s.model, 
                    COALESCE(s.kolicina, 1) as kolicina,
                    COUNT(DISTINCT rs.radnik_id) as dodeljeno,
                    s.status, 
                    s.napomena,
                    s.driver_link
                FROM stampaci s
                LEFT JOIN radnik_stampaci rs ON s.id = rs.stampac_id
                GROUP BY s.id
                ORDER BY s.model
            """)
        else:
            cursor.execute("""
                SELECT 
                    s.id, 
                    s.model, 
                    COALESCE(s.kolicina, 1) as kolicina,
                    COUNT(DISTINCT rs.radnik_id) as dodeljeno,
                    s.status, 
                    s.napomena,
                    s.driver_link
                FROM stampaci s
                LEFT JOIN radnik_stampaci rs ON s.id = rs.stampac_id
                WHERE s.status = ?
                GROUP BY s.id
                ORDER BY s.model
            """, (db_status_filter,))
        
        rows = cursor.fetchall()
        conn.close()
        
        # Privremeno onemogući signal
        try:
            self.stampaci_table.itemChanged.disconnect(self.on_stampac_item_changed)
        except:
            pass
        
        self.stampaci_table.setRowCount(len(rows))
        for i, row in enumerate(rows):
            stampac_id, model, kolicina, dodeljeno, status, napomena, driver_link = row
            slobodno = kolicina - dodeljeno
            
            # Column 0: ID
            item_id = QTableWidgetItem(str(stampac_id))
            item_id.setFlags(item_id.flags() & ~Qt.ItemIsEditable)
            self.stampaci_table.setItem(i, 0, item_id)
            
            # Column 1: Model
            item_model = QTableWidgetItem(model)
            item_model.setFlags(item_model.flags() | Qt.ItemIsEditable)
            self.stampaci_table.setItem(i, 1, item_model)
            
            # Column 2: Količina (editable)
            item_kol = QTableWidgetItem(str(kolicina))
            item_kol.setFlags(item_kol.flags() | Qt.ItemIsEditable)
            self.stampaci_table.setItem(i, 2, item_kol)
            
            # Column 3: Dodeljeno (NOT editable)
            item_dod = QTableWidgetItem(str(dodeljeno))
            item_dod.setFlags(item_dod.flags() & ~Qt.ItemIsEditable)
            self.stampaci_table.setItem(i, 3, item_dod)
            
            # Column 4: Slobodno (NOT editable)
            item_slob = QTableWidgetItem(str(slobodno))
            item_slob.setFlags(item_slob.flags() & ~Qt.ItemIsEditable)
            # Color code: green if available, red if none
            if slobodno > 0:
                item_slob.setBackground(QColor(200, 255, 200))  # Light green
            elif slobodno == 0:
                item_slob.setBackground(QColor(255, 220, 220))  # Light red
            self.stampaci_table.setItem(i, 4, item_slob)
            
            # Column 5: Status (editable, with translation)
            status_map = {
                'Aktivan': T.get('status_active', self.lang),
                'Na servisu': T.get('status_in_service', self.lang),
                'Za rashod': T.get('status_for_disposal', self.lang)
            }
            display_status = status_map.get(status, status)
            item_status = QTableWidgetItem(display_status)
            item_status.setFlags(item_status.flags() | Qt.ItemIsEditable)
            if status == "Na servisu":
                item_status.setBackground(QColor(255, 255, 200))
            elif status == "Za rashod":
                item_status.setBackground(QColor(255, 150, 150))
            self.stampaci_table.setItem(i, 5, item_status)
            
            # Column 6: Napomena (editable)
            item_nap = QTableWidgetItem(napomena if napomena else "")
            item_nap.setFlags(item_nap.flags() | Qt.ItemIsEditable)
            self.stampaci_table.setItem(i, 6, item_nap)
            
            # Column 7: Driver Link (editable)
            item_driver = QTableWidgetItem(driver_link if driver_link else "")
            item_driver.setFlags(item_driver.flags() | Qt.ItemIsEditable)
            # Style as link if has value
            if driver_link and driver_link.strip() and not self.search_active:
                item_driver.setForeground(QColor(0, 0, 255))  # Blue
                font = item_driver.font()
                font.setUnderline(True)
                item_driver.setFont(font)
                item_driver.setToolTip(T.get("tooltip_click_link", self.lang))
            self.stampaci_table.setItem(i, 7, item_driver)
        
        # Ponovo omogući signal
        try:
            self.stampaci_table.itemChanged.disconnect(self.on_stampac_item_changed)
        except:
            pass
        self.stampaci_table.itemChanged.connect(self.on_stampac_item_changed)
        
        # Ažuriraj ukupan broj štampača
        self.update_ukupno_stampaca()
        
        # Re-apply search highlighting if search is active
        if self.search_active:
            self.apply_search_highlighting()
    
    def load_radnici(self):
        conn = self.db.get_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT id, ime, prezime FROM radnici ORDER BY prezime, ime")
        rows = cursor.fetchall()
        conn.close()
        
        # Privremeno onemogući signal
        try:
            self.radnici_table.itemChanged.disconnect(self.on_radnik_item_changed)
        except:
            pass
        
        self.radnici_table.setRowCount(len(rows))
        for i, row in enumerate(rows):
            for j, value in enumerate(row):
                item = QTableWidgetItem(str(value) if value else "")
                
                # SVE kolone osim ID (0) su editabilne
                if j == 0:
                    item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                else:
                    item.setFlags(item.flags() | Qt.ItemIsEditable)
                
                self.radnici_table.setItem(i, j, item)
        
        # Ponovo omogući signal
        try:
            self.radnici_table.itemChanged.disconnect(self.on_radnik_item_changed)
        except:
            pass
        self.radnici_table.itemChanged.connect(self.on_radnik_item_changed)
    
        # Re-apply search highlighting
        if hasattr(self, "search_input") and self.search_input.text():
            self.apply_search_highlighting()

    def search_all(self):
        """Pretražuje sve tabele po unetom tekstu"""
        try:
            search_text = self.search_input.text().lower()
            
            # Ako je search prazan, jednostavno prikaži sve redove
            if not search_text:
                self.search_active = False  # Clear search flag
                
                # Reset tab styles
                try:
                    for i in range(self.tabs.count()):
                        self.tabs.tabBar().setTabTextColor(i, QColor(0, 0, 0))  # Black
                except:
                    pass
                
                # Privremeno blokiraj search signal da ne bi došlo do rekurzije
                try:
                    self.search_input.blockSignals(True)
                    
                    # Pokaži sve redove u svim tabelama
                    if hasattr(self, 'toneri_table'):
                        for i in range(self.toneri_table.rowCount()):
                            self.toneri_table.setRowHidden(i, False)
                    if hasattr(self, 'stampaci_table'):
                        for i in range(self.stampaci_table.rowCount()):
                            self.stampaci_table.setRowHidden(i, False)
                    if hasattr(self, 'radnici_table'):
                        for i in range(self.radnici_table.rowCount()):
                            self.radnici_table.setRowHidden(i, False)
                    if hasattr(self, 'pregled_table'):
                        for i in range(self.pregled_table.rowCount()):
                            self.pregled_table.setRowHidden(i, False)
                    
                    # Resetuj boje tako što ćemo ponovo učitati tabele
                    QTimer.singleShot(100, self.delayed_reload_tables)
                    
                except Exception as e:
                    print(f"Error in search reset: {e}")
                finally:
                    try:
                        self.search_input.blockSignals(False)
                    except:
                        pass
                return
            
            # Disconnect all itemChanged signals to prevent triggering during search
            try:
                self.toneri_table.itemChanged.disconnect(self.on_toner_stanje_changed)
            except:
                pass
            try:
                self.stampaci_table.itemChanged.disconnect(self.on_stampac_item_changed)
            except:
                pass
            try:
                self.radnici_table.itemChanged.disconnect(self.on_radnik_item_changed)
            except:
                pass
            
            # Define highlight color
            self.search_active = True  # Set flag to preserve highlighting
            highlight_color = QColor(34, 139, 34)  # DARK GREEN
            default_color = QColor(255, 255, 255)    # White
            
            # Pretraži tonere
            if hasattr(self, 'toneri_table'):
                for i in range(self.toneri_table.rowCount()):
                    # Check if this toner is below minimum
                    min_qty_item = self.toneri_table.item(i, 2)  # Column 2: Min. količina
                    stock_item = self.toneri_table.item(i, 3)    # Column 3: Stanje
                    is_below_min = False
                    if min_qty_item and stock_item:
                        try:
                            min_qty = int(min_qty_item.text())
                            stock = int(stock_item.text())
                            is_below_min = stock < min_qty
                        except:
                            pass
                    
                    row_has_match = False
                    for j in range(self.toneri_table.columnCount()):
                        item = self.toneri_table.item(i, j)
                        if item:
                            cell_has_match = search_text in item.text().lower()
                            if cell_has_match:
                                row_has_match = True
                            
                            # Reset formatting first
                            try:
                                font = item.font()
                                font.setBold(False)
                                item.setFont(font)
                                item.setForeground(QColor(0, 0, 0))  # Black text
                            except RuntimeError:
                                pass
                            
                            # Apply appropriate background
                            try:
                                if cell_has_match:
                                    # This specific cell matches - GREEN highlight
                                    item.setBackground(highlight_color)
                                    font = item.font()
                                    font.setBold(True)
                                    item.setFont(font)
                                    item.setForeground(QColor(255, 255, 255))  # White text for visibility
                                else:
                                    # No match in this cell - preserve original color
                                    if is_below_min:
                                        item.setBackground(QColor(255, 200, 200))  # Red - below minimum
                                    else:
                                        item.setBackground(default_color)  # White
                            except RuntimeError:
                                pass
                    
                    self.toneri_table.setRowHidden(i, not row_has_match)
            
            # Pretraži štampače
            if hasattr(self, 'stampaci_table'):
                for i in range(self.stampaci_table.rowCount()):
                    # Get Slobodno value for this row (column 4)
                    slobodno_item = self.stampaci_table.item(i, 4)
                    slobodno_value = 0
                    if slobodno_item:
                        try:
                            slobodno_value = int(slobodno_item.text())
                        except:
                            pass
                    
                    row_has_match = False
                    for j in range(self.stampaci_table.columnCount()):
                        item = self.stampaci_table.item(i, j)
                        if item:
                            cell_has_match = search_text in item.text().lower()
                            if cell_has_match:
                                row_has_match = True
                            
                            # Reset formatting first
                            try:
                                font = item.font()
                                font.setBold(False)
                                item.setFont(font)
                                item.setForeground(QColor(0, 0, 0))  # Black text
                            except RuntimeError:
                                pass
                            
                            # Apply appropriate background
                            try:
                                if cell_has_match:
                                    # This specific cell matches - GREEN highlight
                                    item.setBackground(highlight_color)
                                    font = item.font()
                                    font.setBold(True)
                                    item.setFont(font)
                                    item.setForeground(QColor(255, 255, 255))  # White text for visibility
                                else:
                                    # No match in this cell - preserve original colors
                                    if j == 4:  # Slobodno column
                                        if slobodno_value > 0:
                                            item.setBackground(QColor(200, 255, 200))  # Green
                                        elif slobodno_value == 0:
                                            item.setBackground(QColor(255, 220, 220))  # Red
                                        else:
                                            item.setBackground(default_color)
                                    elif j == 5:  # Status column
                                        status_text = item.text()
                                        if T.get("status_in_service", self.lang) in status_text or "Na servisu" in status_text:
                                            item.setBackground(QColor(255, 255, 200))  # Yellow
                                        elif T.get("status_for_disposal", self.lang) in status_text or "Za rashod" in status_text:
                                            item.setBackground(QColor(255, 150, 150))  # Red
                                        else:
                                            item.setBackground(default_color)
                                    else:
                                        item.setBackground(default_color)
                            except RuntimeError:
                                pass
                    
                    self.stampaci_table.setRowHidden(i, not row_has_match)
            
            # Pretraži radnike
            if hasattr(self, 'radnici_table'):
                for i in range(self.radnici_table.rowCount()):
                    row_has_match = False
                    for j in range(self.radnici_table.columnCount()):
                        item = self.radnici_table.item(i, j)
                        if item:
                            cell_has_match = search_text in item.text().lower()
                            if cell_has_match:
                                row_has_match = True
                            
                            # Reset formatting first
                            try:
                                font = item.font()
                                font.setBold(False)
                                item.setFont(font)
                                item.setForeground(QColor(0, 0, 0))  # Black text
                            except RuntimeError:
                                pass
                            
                            # Apply appropriate background
                            try:
                                if cell_has_match:
                                    # This specific cell matches - GREEN highlight
                                    item.setBackground(highlight_color)
                                    font = item.font()
                                    font.setBold(True)
                                    item.setFont(font)
                                    item.setForeground(QColor(255, 255, 255))  # White text for visibility
                                else:
                                    # No match in this cell
                                    item.setBackground(default_color)
                            except RuntimeError:
                                pass
                    self.radnici_table.setRowHidden(i, not row_has_match)
            
            # Pretraži pregled (NO HIGHLIGHTING - just hide/show)
            if hasattr(self, 'pregled_table'):
                for i in range(self.pregled_table.rowCount()):
                    match = False
                    for j in range(self.pregled_table.columnCount()):
                        item = self.pregled_table.item(i, j)
                        if item and search_text in item.text().lower():
                            match = True
                            break
                    self.pregled_table.setRowHidden(i, not match)

            # Highlight tabs that have matches
            # Check each table for matches
            toneri_has_match = False
            stampaci_has_match = False
            radnici_has_match = False
            
            if hasattr(self, 'toneri_table'):
                toneri_has_match = any(not self.toneri_table.isRowHidden(i) for i in range(self.toneri_table.rowCount()))
            
            if hasattr(self, 'stampaci_table'):
                stampaci_has_match = any(not self.stampaci_table.isRowHidden(i) for i in range(self.stampaci_table.rowCount()))
            
            if hasattr(self, 'radnici_table'):
                radnici_has_match = any(not self.radnici_table.isRowHidden(i) for i in range(self.radnici_table.rowCount()))
            
            # Highlight tab buttons (indices: 0=Toneri, 1=Stampaci, 2=Radnici)
            try:
                if toneri_has_match:
                    self.tabs.tabBar().setTabTextColor(0, QColor(34, 139, 34))  # DARK GREEN
                else:
                    self.tabs.tabBar().setTabTextColor(0, QColor(0, 0, 0))  # Black
                
                if stampaci_has_match:
                    self.tabs.tabBar().setTabTextColor(1, QColor(34, 139, 34))  # DARK GREEN
                else:
                    self.tabs.tabBar().setTabTextColor(1, QColor(0, 0, 0))  # Black
                
                if radnici_has_match:
                    self.tabs.tabBar().setTabTextColor(2, QColor(34, 139, 34))  # DARK GREEN
                else:
                    self.tabs.tabBar().setTabTextColor(2, QColor(0, 0, 0))  # Black
            except:
                pass
            
            # Reconnect signals
            try:
                self.toneri_table.itemChanged.connect(self.on_toner_stanje_changed)
            except:
                pass
            try:
                self.stampaci_table.itemChanged.connect(self.on_stampac_item_changed)
            except:
                pass
            try:
                self.radnici_table.itemChanged.connect(self.on_radnik_item_changed)
            except:
                pass
            
        except Exception as e:
            print(f"Critical error in search_all: {e}")
            # Reset everything if there's a critical error
            self.search_active = False
            QTimer.singleShot(100, self.delayed_reload_tables)
    
    def delayed_reload_tables(self):
        """Reload tables with a delay to prevent UI freezing"""
        try:
            # Disconnect signals temporarily
            try:
                self.toneri_table.itemChanged.disconnect(self.on_toner_stanje_changed)
            except:
                pass
            try:
                self.stampaci_table.itemChanged.disconnect(self.on_stampac_item_changed)
            except:
                pass
            try:
                self.radnici_table.itemChanged.disconnect(self.on_radnik_item_changed)
            except:
                pass
            
            # Reload tables
            self.load_toneri()
            self.load_stampaci()
            self.load_radnici()
            self.load_pregled()
            
            # Reconnect signals
            try:
                self.toneri_table.itemChanged.connect(self.on_toner_stanje_changed)
            except:
                pass
            try:
                self.stampaci_table.itemChanged.connect(self.on_stampac_item_changed)
            except:
                pass
            try:
                self.radnici_table.itemChanged.connect(self.on_radnik_item_changed)
            except:
                pass
            
        except Exception as e:
            print(f"Error in delayed_reload_tables: {e}")
    

    def apply_search_highlighting(self):
        """Apply highlighting to search results in ALL tables"""
        try:
            search_text = self.search_input.text().lower()
            if not search_text:
                return
            
            # DISCONNECT all itemChanged signals to prevent recursion
            try:
                self.toneri_table.itemChanged.disconnect(self.on_toner_stanje_changed)
            except:
                pass
            try:
                self.stampaci_table.itemChanged.disconnect(self.on_stampac_item_changed)
            except:
                pass
            try:
                self.radnici_table.itemChanged.disconnect(self.on_radnik_item_changed)
            except:
                pass
            
            highlight_color = QColor(34, 139, 34)  # DARK GREEN
            
            # Highlight in TONERI table
            if hasattr(self, 'toneri_table'):
                for i in range(self.toneri_table.rowCount()):
                    if not self.toneri_table.isRowHidden(i):
                        for j in range(self.toneri_table.columnCount()):
                            item = self.toneri_table.item(i, j)
                            if item and search_text in item.text().lower():
                                try:
                                    item.setBackground(highlight_color)
                                    font = item.font()
                                    font.setBold(True)
                                    item.setFont(font)
                                    item.setForeground(QColor(255, 255, 255))  # White text
                                except RuntimeError:
                                    pass
            
            # Highlight in STAMPACI table
            if hasattr(self, 'stampaci_table'):
                for i in range(self.stampaci_table.rowCount()):
                    if not self.stampaci_table.isRowHidden(i):
                        for j in range(self.stampaci_table.columnCount()):
                            item = self.stampaci_table.item(i, j)
                            if item and search_text in item.text().lower():
                                try:
                                    item.setBackground(highlight_color)
                                    font = item.font()
                                    font.setBold(True)
                                    item.setFont(font)
                                    item.setForeground(QColor(255, 255, 255))  # White text
                                except RuntimeError:
                                    pass
            
            # Highlight in RADNICI table
            if hasattr(self, 'radnici_table'):
                for i in range(self.radnici_table.rowCount()):
                    if not self.radnici_table.isRowHidden(i):
                        for j in range(self.radnici_table.columnCount()):
                            item = self.radnici_table.item(i, j)
                            if item and search_text in item.text().lower():
                                try:
                                    item.setBackground(highlight_color)
                                    font = item.font()
                                    font.setBold(True)
                                    item.setFont(font)
                                    item.setForeground(QColor(255, 255, 255))  # White text
                                except RuntimeError:
                                    pass
            
            # PREGLED table - NO highlighting, just hide/show handled by search_all
            
            # RECONNECT all signals
            try:
                self.toneri_table.itemChanged.connect(self.on_toner_stanje_changed)
            except:
                pass
            try:
                self.stampaci_table.itemChanged.connect(self.on_stampac_item_changed)
            except:
                pass
            try:
                self.radnici_table.itemChanged.connect(self.on_radnik_item_changed)
            except:
                pass
            
        except Exception as e:
            print(f"Error in apply_search_highlighting: {e}")
    
    def add_toner(self):
        dialog = TonerDialog(self)
        if dialog.exec_():
            data = dialog.get_data()
            if not data['model']:
                QMessageBox.warning(self, T.get("error", self.lang), T.get("error_model_required", self.lang))
                return
            
            try:
                conn = self.db.get_connection()
                cursor = conn.cursor()
                cursor.execute('''
                    INSERT INTO toneri (model, minimalna_kolicina, trenutno_stanje)
                    VALUES (?, ?, ?)
                ''', (data['model'], data['minimalna_kolicina'], 
                      data['trenutno_stanje']))
                conn.commit()
                conn.close()
                self.load_toneri()
                QMessageBox.information(self, T.get("success", self.lang), T.get("msg_toner_added", self.lang))
            except sqlite3.IntegrityError:
                QMessageBox.warning(self, T.get("error", self.lang), T.get("error_model_exists", self.lang))
    
    def edit_toner(self):
        selected = self.toneri_table.selectedItems()
        if not selected:
            QMessageBox.warning(self, T.get("error", self.lang), T.get("error_select_toner", self.lang))
            return
        
        row = selected[0].row()
        toner_id = int(self.toneri_table.item(row, 0).text())
        
        conn = self.db.get_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM toneri WHERE id = ?", (toner_id,))
        toner_data = cursor.fetchone()
        conn.close()
        
        dialog = TonerDialog(self, toner_data)
        if dialog.exec_():
            data = dialog.get_data()
            try:
                conn = self.db.get_connection()
                cursor = conn.cursor()
                cursor.execute('''
                    UPDATE toneri SET model=?, minimalna_kolicina=?, 
                    trenutno_stanje=? WHERE id=?
                ''', (data['model'], data['minimalna_kolicina'],
                      data['trenutno_stanje'], toner_id))
                conn.commit()
                conn.close()
                self.load_toneri()
                QMessageBox.information(self, T.get("success", self.lang), T.get("msg_toner_edited", self.lang))
            except sqlite3.IntegrityError:
                QMessageBox.warning(self, T.get("error", self.lang), "Model već postoji!")
    
    def delete_toner(self):
        selected = self.toneri_table.selectedItems()
        if not selected:
            QMessageBox.warning(self, T.get("error", self.lang), T.get("error_select_toner_delete", self.lang))
            return
        
        row = selected[0].row()
        toner_id = int(self.toneri_table.item(row, 0).text())
        model = self.toneri_table.item(row, 1).text()
        
        # Check if toner is linked to any printers
        conn = self.db.get_connection()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT s.model 
            FROM stampaci s
            JOIN stampac_toneri st ON s.id = st.stampac_id
            WHERE st.toner_id = ?
        """, (toner_id,))
        linked_printers = cursor.fetchall()
        
        if linked_printers:
            printers_list = ", ".join([model for (model,) in linked_printers])
            reply = QMessageBox.question(
                self, 
                T.get("confirm", self.lang),
                T.get("toner_linked_to", self.lang).format(model) + "\n" + printers_list + "\n\n" +
                T.get("delete_will_unlink", self.lang) + "\n\n" +
                T.get("confirm_delete_toner", self.lang),
                QMessageBox.Yes | QMessageBox.No
            )
            if reply != QMessageBox.Yes:
                conn.close()
                return
        else:
            reply = QMessageBox.question(self, T.get("confirm", self.lang), T.get("confirm_delete_toner", self.lang),
                                         QMessageBox.Yes | QMessageBox.No)
            if reply != QMessageBox.Yes:
                conn.close()
                return
        
        # Delete all relationships first
        cursor.execute("DELETE FROM stampac_toneri WHERE toner_id = ?", (toner_id,))
        cursor.execute("DELETE FROM istorija_narudzbi WHERE toner_id = ?", (toner_id,))
        # Then delete the toner
        cursor.execute("DELETE FROM toneri WHERE id = ?", (toner_id,))
        conn.commit()
        conn.close()
        
        self.load_toneri()
        self.load_stampaci()  # Refresh printer list too
        self.load_pregled()  # Refresh overview
        QMessageBox.information(self, T.get("success", self.lang), T.get("toner_deleted", self.lang))
    
    def add_stampac(self):
        dialog = StampacDialog(self, self.db)
        if dialog.exec_():
            data = dialog.get_data()
            if not data['model']:
                QMessageBox.warning(self, T.get("error", self.lang), T.get("error_printer_model_required", self.lang))
                return
            
            try:
                conn = self.db.get_connection()
                cursor = conn.cursor()
                cursor.execute('''
                    INSERT INTO stampaci (model, kolicina, status, napomena, driver_link)
                    VALUES (?, ?, ?, ?, ?)
                ''', (data['model'], data['kolicina'], data['status'], data['napomena'], data['driver_link']))
                
                stampac_id = cursor.lastrowid
                
                # Sačuvaj veze sa tonerima
                selected_toneri = dialog.get_selected_toneri()
                for toner_id in selected_toneri:
                    cursor.execute('''
                        INSERT INTO stampac_toneri (stampac_id, toner_id)
                        VALUES (?, ?)
                    ''', (stampac_id, toner_id))
                
                conn.commit()
                conn.close()
                self.load_stampaci()
                self.load_pregled()  # Osvježi pregled tab
                QMessageBox.information(self, T.get("success", self.lang), T.get("msg_printer_added", self.lang))
            except sqlite3.IntegrityError:
                QMessageBox.warning(self, T.get("error", self.lang), "Serijski broj već postoji!")
    
    def edit_stampac(self):
        selected = self.stampaci_table.selectedItems()
        if not selected:
            QMessageBox.warning(self, T.get("error", self.lang), T.get("error_select_printer", self.lang))
            return
        
        row = selected[0].row()
        stampac_id = int(self.stampaci_table.item(row, 0).text())
        
        conn = self.db.get_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM stampaci WHERE id = ?", (stampac_id,))
        stampac_data = cursor.fetchone()
        conn.close()
        
        dialog = StampacDialog(self, self.db, stampac_data)
        if dialog.exec_():
            data = dialog.get_data()
            try:
                conn = self.db.get_connection()
                cursor = conn.cursor()
                cursor.execute('''
                    UPDATE stampaci SET model=?, kolicina=?, status=?, napomena=?, driver_link=? WHERE id=?
                ''', (data['model'], data['kolicina'], data['status'], data['napomena'], data['driver_link'], stampac_id))
                
                # Obriši stare veze sa tonerima
                cursor.execute("DELETE FROM stampac_toneri WHERE stampac_id = ?", (stampac_id,))
                
                # Sačuvaj nove veze
                selected_toneri = dialog.get_selected_toneri()
                for toner_id in selected_toneri:
                    cursor.execute('''
                        INSERT INTO stampac_toneri (stampac_id, toner_id)
                        VALUES (?, ?)
                    ''', (stampac_id, toner_id))
                
                conn.commit()
                conn.close()
                self.load_stampaci()
                self.load_pregled()  # Osvježi pregled tab
                QMessageBox.information(self, T.get("success", self.lang), T.get("msg_printer_edited", self.lang))
            except sqlite3.IntegrityError:
                QMessageBox.warning(self, T.get("error", self.lang), "Serijski broj već postoji!")
    
    def delete_stampac(self):
        selected = self.stampaci_table.selectedItems()
        if not selected:
            QMessageBox.warning(self, T.get("error", self.lang), T.get("error_select_printer_delete", self.lang))
            return
        
        row = selected[0].row()
        stampac_id = int(self.stampaci_table.item(row, 0).text())
        model = self.stampaci_table.item(row, 1).text()
        
        # Check if printer is assigned to any employee
        conn = self.db.get_connection()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT r.ime, r.prezime 
            FROM radnici r
            JOIN radnik_stampaci rs ON r.id = rs.radnik_id
            WHERE rs.stampac_id = ?
        """, (stampac_id,))
        assigned_employees = cursor.fetchall()
        
        # Check if printer is linked to any toners
        cursor.execute("""
            SELECT t.model 
            FROM toneri t
            JOIN stampac_toneri st ON t.id = st.toner_id
            WHERE st.stampac_id = ?
        """, (stampac_id,))
        linked_toners = cursor.fetchall()
        
        # Build confirmation message
        if assigned_employees or linked_toners:
            message_parts = []
            
            if assigned_employees:
                employees_list = "\n".join([f"• {ime} {prezime}" for ime, prezime in assigned_employees])
                if self.lang == 'sr':
                    message_parts.append(f"Dodeljeno radnicima:\n{employees_list}")
                else:
                    message_parts.append(f"Assigned to employees:\n{employees_list}")
            
            if linked_toners:
                toners_list = "\n".join([f"• {model}" for (model,) in linked_toners])
                if self.lang == 'sr':
                    message_parts.append(f"Koristi tonere:\n{toners_list}")
                else:
                    message_parts.append(f"Uses toners:\n{toners_list}")
            
            full_message = (T.get("printer_assigned_to", self.lang).format(model) if assigned_employees 
                          else f"Štampač '{model}' ima veze:" if self.lang == 'sr' 
                          else f"Printer '{model}' has connections:")
            full_message += "\n\n" + "\n\n".join(message_parts) + "\n\n"
            full_message += T.get("delete_will_unassign", self.lang) + "\n\n"
            full_message += T.get("confirm_delete_printer", self.lang)
            
            reply = QMessageBox.question(
                self, 
                T.get("confirm", self.lang),
                full_message,
                QMessageBox.Yes | QMessageBox.No
            )
            if reply != QMessageBox.Yes:
                conn.close()
                return
        else:
            reply = QMessageBox.question(self, T.get("confirm", self.lang), T.get("confirm_delete_printer", self.lang),
                                         QMessageBox.Yes | QMessageBox.No)
            if reply != QMessageBox.Yes:
                conn.close()
                return
        
        # Delete all relationships first
        cursor.execute("DELETE FROM radnik_stampaci WHERE stampac_id = ?", (stampac_id,))
        cursor.execute("DELETE FROM stampac_toneri WHERE stampac_id = ?", (stampac_id,))
        # Then delete the printer
        cursor.execute("DELETE FROM stampaci WHERE id = ?", (stampac_id,))
        conn.commit()
        conn.close()
        
        self.load_stampaci()
        self.load_radnici()  # Refresh employee list too
        self.load_pregled()  # Refresh overview
        QMessageBox.information(self, T.get("success", self.lang), T.get("printer_deleted", self.lang))
    
    def add_radnik(self):
        dialog = RadnikDialog(self, self.db)
        if dialog.exec_():
            data = dialog.get_data()
            if not data['ime'] or not data['prezime']:
                QMessageBox.warning(self, T.get("error", self.lang), T.get("error_name_required", self.lang))
                return
            
            conn = self.db.get_connection()
            cursor = conn.cursor()
            cursor.execute('''
                INSERT INTO radnici (ime, prezime)
                VALUES (?, ?)
            ''', (data['ime'], data['prezime']))
            
            radnik_id = cursor.lastrowid
            
            # Sačuvaj veze sa štampačima - ali prvo proveri dostupnost!
            selected_stampaci = dialog.get_selected_stampaci()
            
            # Check availability for each printer
            for stampac_id in selected_stampaci:
                cursor.execute("""
                    SELECT 
                        s.model,
                        COALESCE(s.kolicina, 1) as kolicina,
                        COUNT(DISTINCT rs.radnik_id) as dodeljeno
                    FROM stampaci s
                    LEFT JOIN radnik_stampaci rs ON s.id = rs.stampac_id
                    WHERE s.id = ?
                    GROUP BY s.id
                """, (stampac_id,))
                result = cursor.fetchone()
                
                if result:
                    model, kolicina, dodeljeno = result
                    slobodno = kolicina - dodeljeno
                    
                    if slobodno <= 0:
                        conn.rollback()
                        conn.close()
                        QMessageBox.warning(self, T.get("error", self.lang), 
                            f"Štampač '{model}' nema slobodnih jedinica!\n\nUkupno: {kolicina}\nDodeljeno: {dodeljeno}\nSlobodno: 0")
                        return
            
            # All checks passed, insert assignments
            for stampac_id in selected_stampaci:
                cursor.execute('''
                    INSERT INTO radnik_stampaci (radnik_id, stampac_id)
                    VALUES (?, ?)
                ''', (radnik_id, stampac_id))
            
            conn.commit()
            conn.close()
            self.load_radnici()
            self.load_pregled()  # Osvježi pregled tab
            QMessageBox.information(self, T.get("success", self.lang), T.get("msg_employee_added", self.lang))
    
    def edit_radnik(self):
        selected = self.radnici_table.selectedItems()
        if not selected:
            QMessageBox.warning(self, T.get("error", self.lang), T.get("error_select_employee", self.lang))
            return
        
        row = selected[0].row()
        radnik_id = int(self.radnici_table.item(row, 0).text())
        
        conn = self.db.get_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM radnici WHERE id = ?", (radnik_id,))
        radnik_data = cursor.fetchone()
        conn.close()
        
        dialog = RadnikDialog(self, self.db, radnik_data)
        if dialog.exec_():
            data = dialog.get_data()
            conn = self.db.get_connection()
            cursor = conn.cursor()
            cursor.execute('''
                UPDATE radnici SET ime=?, prezime=? WHERE id=?
            ''', (data['ime'], data['prezime'], radnik_id))
            
            # Obriši stare veze sa štampačima
            cursor.execute("SELECT stampac_id FROM radnik_stampaci WHERE radnik_id = ?", (radnik_id,))
            old_stampaci = {row[0] for row in cursor.fetchall()}
            cursor.execute("DELETE FROM radnik_stampaci WHERE radnik_id = ?", (radnik_id,))
            
            # Sačuvaj nove veze - proveri dostupnost za nove dodeljene
            selected_stampaci = dialog.get_selected_stampaci()
            new_stampaci = set(selected_stampaci) - old_stampaci  # Only new assignments
            
            # Check availability only for NEW assignments
            for stampac_id in new_stampaci:
                cursor.execute("""
                    SELECT 
                        s.model,
                        COALESCE(s.kolicina, 1) as kolicina,
                        COUNT(DISTINCT rs.radnik_id) as dodeljeno
                    FROM stampaci s
                    LEFT JOIN radnik_stampaci rs ON s.id = rs.stampac_id
                    WHERE s.id = ?
                    GROUP BY s.id
                """, (stampac_id,))
                result = cursor.fetchone()
                
                if result:
                    model, kolicina, dodeljeno = result
                    slobodno = kolicina - dodeljeno
                    
                    if slobodno <= 0:
                        conn.rollback()
                        conn.close()
                        QMessageBox.warning(self, T.get("error", self.lang), 
                            f"Štampač '{model}' nema slobodnih jedinica!\n\nUkupno: {kolicina}\nDodeljeno: {dodeljeno}\nSlobodno: 0")
                        return
            
            # All checks passed, insert all selected
            for stampac_id in selected_stampaci:
                cursor.execute('''
                    INSERT INTO radnik_stampaci (radnik_id, stampac_id)
                    VALUES (?, ?)
                ''', (radnik_id, stampac_id))
            
            conn.commit()
            conn.close()
            self.load_radnici()
            self.load_pregled()  # Osvježi pregled tab
            QMessageBox.information(self, T.get("success", self.lang), T.get("msg_employee_edited", self.lang))
    
    def delete_radnik(self):
        selected = self.radnici_table.selectedItems()
        if not selected:
            QMessageBox.warning(self, T.get("error", self.lang), T.get("error_select_employee_delete", self.lang))
            return
        
        row = selected[0].row()
        radnik_id = int(self.radnici_table.item(row, 0).text())
        
        reply = QMessageBox.question(self, T.get("confirm", self.lang), T.get("confirm_delete_employee", self.lang),
                                     QMessageBox.Yes | QMessageBox.No)
        if reply == QMessageBox.Yes:
            conn = self.db.get_connection()
            cursor = conn.cursor()
            cursor.execute("DELETE FROM radnici WHERE id = ?", (radnik_id,))
            conn.commit()
            conn.close()
            self.load_radnici()
    
    def evidentira_potrosnju(self):
        """Dialog za evidentiranje potrošnje tonera"""
        selected = self.toneri_table.selectedItems()
        if not selected:
            QMessageBox.warning(self, T.get("error", self.lang), T.get("error_select_toner_consumption", self.lang))
            return
        
        row = selected[0].row()
        toner_id = int(self.toneri_table.item(row, 0).text())
        trenutno_stanje = int(self.toneri_table.item(row, 3).text())  # Kolona 3 je Stanje (bez Opisa)
        
        if trenutno_stanje <= 0:
            QMessageBox.warning(self, T.get("error", self.lang), T.get("error_no_stock", self.lang))
            return
        
        reply = QMessageBox.question(self, T.get("confirm", self.lang), 
                                     T.get("confirm_reduce_stock", self.lang).format(trenutno_stanje),
                                     QMessageBox.Yes | QMessageBox.No)
        if reply == QMessageBox.Yes:
            conn = self.db.get_connection()
            cursor = conn.cursor()
            cursor.execute("UPDATE toneri SET trenutno_stanje = trenutno_stanje - 1 WHERE id = ?", (toner_id,))
            
            # Dodaj u istoriju potrošnje
            cursor.execute("INSERT INTO istorija_potrosnje (toner_id) VALUES (?)", (toner_id,))
            
            conn.commit()
            conn.close()
            self.load_toneri()
            QMessageBox.information(self, T.get("success", self.lang), T.get("msg_consumption_recorded", self.lang))
    
    def prikazi_narudzbu(self):
        """Prikazuje sve tonere gde je stanje < minimalna količina"""
        conn = self.db.get_connection()
        cursor = conn.cursor()
        
        # DEBUG: Prvo proveri sve tonere
        cursor.execute("SELECT model, minimalna_kolicina, trenutno_stanje FROM toneri")
        debug_rows = cursor.fetchall()
        debug_info = []
        for row in debug_rows:
            model, min_kol, stanje = row
            debug_info.append(f"{model}: min={min_kol} (tip: {type(min_kol)}), stanje={stanje} (tip: {type(stanje)})")
        
        cursor.execute('''
            SELECT model, minimalna_kolicina, trenutno_stanje, 
                   (minimalna_kolicina - trenutno_stanje) as za_narucivanje
            FROM toneri 
            WHERE CAST(trenutno_stanje AS INTEGER) < CAST(minimalna_kolicina AS INTEGER)
            ORDER BY za_narucivanje DESC
        ''')
        rows = cursor.fetchall()
        conn.close()
        
        if not rows:
            debug_text = "\n".join(debug_info)
            QMessageBox.information(self, T.get("order_title", self.lang), 
                T.get("order_no_toners", self.lang) + f"\n\n{T.get('order_debug_info', self.lang)}\n{debug_text}")
            return
        
        # Kreiraj dijalog sa listom
        dialog = QDialog(self)
        dialog.setWindowTitle(T.get("order_list_title", self.lang))
        dialog.setMinimumSize(900, 400)
        layout = QVBoxLayout(dialog)
        
        info_label = QLabel(T.get("order_found", self.lang).format(len(rows)))
        info_label.setStyleSheet("font-weight: bold; font-size: 12pt; padding: 10px;")
        layout.addWidget(info_label)
        
        table = QTableWidget()
        table.setColumnCount(4)
        table.setHorizontalHeaderLabels([
            T.get("col_model", self.lang), 
            T.get("col_min_qty", self.lang), 
            T.get("col_current_stock", self.lang), 
            T.get("col_for_order", self.lang)
        ])
        table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        table.horizontalHeader().setStretchLastSection(True)
        table.setRowCount(len(rows))
        
        for i, row in enumerate(rows):
            for j, value in enumerate(row):
                # Prikaži 0 ako je vrednost 0, ne prazan string
                item = QTableWidgetItem(str(value) if value is not None else "")
                if j == 3:  # Za naručivanje kolona (sada na poziciji 3)
                    item.setBackground(QColor(255, 200, 200))
                    item.setForeground(QColor(200, 0, 0))
                    font = item.font()
                    font.setBold(True)
                    item.setFont(font)
                table.setItem(i, j, item)
        
        layout.addWidget(table)
        
        # Dugmad
        btn_layout = QHBoxLayout()
        preview_btn = QPushButton(T.get("btn_preview_print", self.lang))
        preview_btn.clicked.connect(lambda: self.preview_narudzbu(rows))
        excel_btn = QPushButton(T.get("btn_excel_export", self.lang))
        excel_btn.clicked.connect(lambda: self.export_narudzbu_excel(rows))
        history_btn = QPushButton(T.get("btn_add_to_history", self.lang))
        history_btn.clicked.connect(lambda: self.dodaj_narudzbu_dialog(dialog, rows))
        close_btn = QPushButton(T.get("btn_close", self.lang))
        close_btn.clicked.connect(dialog.close)
        btn_layout.addWidget(preview_btn)
        btn_layout.addWidget(excel_btn)
        btn_layout.addWidget(history_btn)
        btn_layout.addWidget(close_btn)
        layout.addLayout(btn_layout)
        
        dialog.exec_()
    
    def stampaj_tonere(self):
        """Prikazuje preview svih tonera u browseru sa mogućnošću štampanja"""
        try:
            from datetime import datetime
            import tempfile
            
            conn = self.db.get_connection()
            cursor = conn.cursor()
            cursor.execute("""
                SELECT model, minimalna_kolicina, trenutno_stanje 
                FROM toneri 
                ORDER BY model
            """)
            rows = cursor.fetchall()
            
            # Izračunaj ukupan zbir svih tonera
            cursor.execute("SELECT COALESCE(SUM(trenutno_stanje), 0) FROM toneri")
            ukupan_zbir = cursor.fetchone()[0]
            
            conn.close()
            
            if not rows:
                QMessageBox.information(self, T.get("info", self.lang), "Nema tonera u bazi.")
                return
            
            # Kreiraj HTML
            html = f"""
            <!DOCTYPE html>
            <html>
            <head>
                <meta charset="UTF-8">
                <title>{T.get('col_toners', self.lang) if self.lang == 'sr' else 'Toners'}</title>
                <style>
                    body {{ 
                        font-family: Arial, sans-serif; 
                        margin: 40px;
                        max-width: 1200px;
                    }}
                    h1 {{ 
                        text-align: center; 
                        color: #2C3E50; 
                        margin-bottom: 10px;
                        font-size: 24pt;
                    }}
                    .datum {{ 
                        text-align: left; 
                        margin-bottom: 30px;
                        font-size: 12pt;
                    }}
                    table {{ 
                        border-collapse: collapse; 
                        width: 100%;
                        box-shadow: 0 2px 4px rgba(0,0,0,0.1);
                    }}
                    th {{ 
                        background-color: #34495E; 
                        color: white; 
                        padding: 12px; 
                        border: 1px solid #2C3E50;
                        font-size: 12pt;
                    }}
                    td {{ 
                        padding: 10px; 
                        border: 1px solid #BDC3C7; 
                        text-align: center;
                        font-size: 11pt;
                    }}
                    tr:nth-child(even) {{ 
                        background-color: #ECF0F1; 
                    }}
                    tr:hover {{
                        background-color: #D5DBDB;
                    }}
                    .red {{ 
                        background-color: #FFB6B6; 
                        color: #C80000; 
                        font-weight: bold;
                    }}
                    .footer {{ 
                        margin-top: 20px; 
                        font-style: italic;
                        font-size: 11pt;
                        text-align: center;
                    }}
                    .print-button {{
                        background-color: #2196F3;
                        color: white;
                        padding: 15px 32px;
                        text-align: center;
                        font-size: 16px;
                        margin: 20px 0;
                        cursor: pointer;
                        border: none;
                        border-radius: 4px;
                        display: block;
                        margin-left: auto;
                        margin-right: auto;
                    }}
                    .print-button:hover {{
                        background-color: #0b7dda;
                    }}
                    @media print {{
                        .print-button {{ display: none; }}
                        body {{ margin: 20px; }}
                    }}
                </style>
            </head>
            <body>
                <button class="print-button" onclick="window.print()">🖨️ {T.get('preview_print_btn', self.lang)}</button>
                
                <h1>{'LISTA TONERA' if self.lang == 'sr' else 'TONER LIST'}</h1>
                <p class="datum">{'Datum:' if self.lang == 'sr' else 'Date:'} {datetime.now().strftime('%d.%m.%Y.')}</p>
                <table>
                    <tr>
                        <th>{'Redni broj' if self.lang == 'sr' else 'No.'}</th>
                        <th>{T.get('col_model', self.lang)}</th>
                        <th>{T.get('col_min_qty', self.lang)}</th>
                        <th>{T.get('col_stock', self.lang)}</th>
                    </tr>
            """
            
            for idx, row in enumerate(rows, 1):
                model, min_kol, trenutno = row
                # Proveri da li je ispod minimuma
                is_below_min = trenutno < min_kol if trenutno is not None and min_kol is not None else False
                stanje_class = 'class="red"' if is_below_min else ''
                
                html += f"""
                    <tr>
                        <td>{idx}</td>
                        <td>{model}</td>
                        <td>{min_kol if min_kol is not None else '-'}</td>
                        <td {stanje_class}>{trenutno if trenutno is not None else '0'}</td>
                    </tr>
                """
            
            html += f"""
                </table>
                <p class="footer">
                    {'Različitih tonera:' if self.lang == 'sr' else 'Different toners:'} {len(rows)} | 
                    {'Ukupno komada:' if self.lang == 'sr' else 'Total pieces:'} {ukupan_zbir}
                </p>
            </body>
            </html>
            """
            
            # Sačuvaj u temp fajl i otvori u browseru
            temp_file = tempfile.NamedTemporaryFile(mode='w', delete=False, suffix='.html', encoding='utf-8')
            temp_file.write(html)
            temp_file.close()
            
            QDesktopServices.openUrl(QUrl.fromLocalFile(temp_file.name))
            
        except Exception as e:
            QMessageBox.critical(self, T.get("error", self.lang), f"{T.get('error', self.lang)}:\n{str(e)}")
    
    def export_tonere_excel(self):
        """Eksportuje sve tonere u Excel"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from datetime import datetime
            import os
            
            conn = self.db.get_connection()
            cursor = conn.cursor()
            cursor.execute("""
                SELECT model, minimalna_kolicina, trenutno_stanje 
                FROM toneri 
                ORDER BY model
            """)
            rows = cursor.fetchall()
            
            # Izračunaj ukupan zbir svih tonera
            cursor.execute("SELECT COALESCE(SUM(trenutno_stanje), 0) FROM toneri")
            ukupan_zbir = cursor.fetchone()[0]
            
            conn.close()
            
            if not rows:
                QMessageBox.information(self, T.get("info", self.lang), "Nema tonera u bazi.")
                return
            
            wb = Workbook()
            ws = wb.active
            ws.title = 'Lista tonera'[:31]  # Excel limit 31 chars
            
            # Naslov
            ws['A1'] = 'LISTA TONERA' if self.lang == 'sr' else 'TONER LIST'
            ws['A1'].font = Font(size=18, bold=True)
            ws['A1'].alignment = Alignment(horizontal='center')
            ws.merge_cells('A1:D1')
            
            # Datum
            ws['A2'] = f"{'Datum:' if self.lang == 'sr' else 'Date:'} {datetime.now().strftime('%d.%m.%Y.')}"
            ws['A2'].font = Font(size=11)
            ws['A2'].alignment = Alignment(horizontal='left')
            
            # Header
            headers = [
                'Br.' if self.lang == 'sr' else 'No.',
                T.get('col_model', self.lang), 
                T.get('col_min_qty', self.lang), 
                T.get('col_stock', self.lang)
            ]
            
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col)
                cell.value = header
                cell.font = Font(bold=True, color='FFFFFF', size=12)
                cell.fill = PatternFill(start_color='34495E', end_color='34495E', fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
            
            # Data
            for idx, row_data in enumerate(rows, 1):
                model, min_kol, trenutno = row_data
                
                # Proveri da li je ispod minimuma
                is_below_min = trenutno < min_kol if trenutno is not None and min_kol is not None else False
                
                # Redni broj
                cell = ws.cell(row=idx+4, column=1)
                cell.value = idx
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
                
                # Model
                cell = ws.cell(row=idx+4, column=2)
                cell.value = model
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
                
                # Minimalna količina
                cell = ws.cell(row=idx+4, column=3)
                cell.value = min_kol if min_kol is not None else '-'
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
                
                # Trenutno stanje
                cell = ws.cell(row=idx+4, column=4)
                cell.value = trenutno if trenutno is not None else 0
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
                
                # Highlight ako je ispod minimuma
                if is_below_min:
                    cell.fill = PatternFill(start_color='FFB6B6', end_color='FFB6B6', fill_type='solid')
                    cell.font = Font(bold=True, color='C80000')
            
            # Auto width
            for col_idx in range(1, 5):  # 4 kolone
                max_length = 0
                for row in ws.iter_rows(min_row=4, min_col=col_idx, max_col=col_idx):
                    for cell in row:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                from openpyxl.utils import get_column_letter
                ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 3
            
            # Footer
            footer_row = len(rows) + 6
            if self.lang == 'sr':
                footer_text = f"Različitih tonera: {len(rows)} | Ukupno komada: {ukupan_zbir}"
            else:
                footer_text = f"Different toners: {len(rows)} | Total pieces: {ukupan_zbir}"
            ws.cell(row=footer_row, column=1).value = footer_text
            ws.cell(row=footer_row, column=1).font = Font(italic=True, size=11)
            
            # Save dialog
            file_path, _ = QFileDialog.getSaveFileName(
                self, 
                T.get("btn_excel_export", self.lang), 
                f"Lista_tonera_{datetime.now().strftime('%Y%m%d')}.xlsx",
                "Excel Files (*.xlsx)"
            )
            
            if file_path:
                wb.save(file_path)
                QMessageBox.information(self, T.get("success", self.lang), 
                    f"{'Fajl je sačuvan:' if self.lang == 'sr' else 'File saved:'}\n{file_path}")
                
                # Pitaj korisnika da li želi da otvori fajl
                reply = QMessageBox.question(self, 
                    T.get("success", self.lang),
                    T.get("msg_open_file", self.lang) if self.lang == 'sr' else "Open file?",
                    QMessageBox.Yes | QMessageBox.No)
                
                if reply == QMessageBox.Yes:
                    QDesktopServices.openUrl(QUrl.fromLocalFile(file_path))
            
        except Exception as e:
            QMessageBox.critical(self, T.get("error", self.lang), f"{T.get('error', self.lang)}:\n{str(e)}")
    
    def stampaj_stampace(self):
        """Prikazuje preview svih štampača u browseru sa mogućnošću štampanja"""
        try:
            from datetime import datetime
            import tempfile
            
            conn = self.db.get_connection()
            cursor = conn.cursor()
            cursor.execute("""
                SELECT 
                    s.model, 
                    COALESCE(s.kolicina, 1) as kolicina,
                    COUNT(DISTINCT rs.radnik_id) as dodeljeno,
                    s.status,
                    s.napomena
                FROM stampaci s
                LEFT JOIN radnik_stampaci rs ON s.id = rs.stampac_id
                GROUP BY s.id
                ORDER BY s.model
            """)
            rows = cursor.fetchall()
            
            # Izračunaj ukupan broj štampača
            cursor.execute("SELECT COALESCE(SUM(kolicina), 0) FROM stampaci")
            ukupan_broj = cursor.fetchone()[0]
            
            conn.close()
            
            if not rows:
                QMessageBox.information(self, T.get("info", self.lang), "Nema štampača u bazi.")
                return
            
            # Kreiraj HTML
            html = f"""
            <!DOCTYPE html>
            <html>
            <head>
                <meta charset="UTF-8">
                <title>{'LISTA ŠTAMPAČA' if self.lang == 'sr' else 'PRINTER LIST'}</title>
                <style>
                    body {{ 
                        font-family: Arial, sans-serif; 
                        margin: 40px;
                        max-width: 1400px;
                    }}
                    h1 {{ 
                        text-align: center; 
                        color: #2C3E50; 
                        margin-bottom: 10px;
                        font-size: 24pt;
                    }}
                    .datum {{ 
                        text-align: left; 
                        margin-bottom: 30px;
                        font-size: 12pt;
                    }}
                    table {{ 
                        border-collapse: collapse; 
                        width: 100%;
                        box-shadow: 0 2px 4px rgba(0,0,0,0.1);
                    }}
                    th {{ 
                        background-color: #34495E; 
                        color: white; 
                        padding: 12px; 
                        border: 1px solid #2C3E50;
                        font-size: 11pt;
                    }}
                    td {{ 
                        padding: 10px; 
                        border: 1px solid #BDC3C7; 
                        text-align: center;
                        font-size: 10pt;
                    }}
                    tr:nth-child(even) {{ 
                        background-color: #ECF0F1; 
                    }}
                    tr:hover {{
                        background-color: #D5DBDB;
                    }}
                    .green {{ 
                        background-color: #C8E6C9; 
                        color: #2E7D32; 
                        font-weight: bold;
                    }}
                    .red {{ 
                        background-color: #FFCDD2; 
                        color: #C62828; 
                        font-weight: bold;
                    }}
                    .yellow {{ 
                        background-color: #FFF9C4; 
                        color: #F57F17; 
                        font-weight: bold;
                    }}
                    .footer {{ 
                        margin-top: 20px; 
                        font-style: italic;
                        font-size: 11pt;
                        text-align: center;
                    }}
                    .print-button {{
                        background-color: #2196F3;
                        color: white;
                        padding: 15px 32px;
                        text-align: center;
                        font-size: 16px;
                        margin: 20px 0;
                        cursor: pointer;
                        border: none;
                        border-radius: 4px;
                        display: block;
                        margin-left: auto;
                        margin-right: auto;
                    }}
                    .print-button:hover {{
                        background-color: #0b7dda;
                    }}
                    @media print {{
                        .print-button {{ display: none; }}
                        body {{ margin: 20px; }}
                    }}
                </style>
            </head>
            <body>
                <button class="print-button" onclick="window.print()">🖨️ {T.get('preview_print_btn', self.lang)}</button>
                
                <h1>{'LISTA ŠTAMPAČA' if self.lang == 'sr' else 'PRINTER LIST'}</h1>
                <p class="datum">{'Datum:' if self.lang == 'sr' else 'Date:'} {datetime.now().strftime('%d.%m.%Y.')}</p>
                <table>
                    <tr>
                        <th>{'Br.' if self.lang == 'sr' else 'No.'}</th>
                        <th>{T.get('col_model', self.lang)}</th>
                        <th>{T.get('col_quantity', self.lang)}</th>
                        <th>{T.get('col_assigned', self.lang)}</th>
                        <th>{T.get('col_available', self.lang)}</th>
                        <th>{T.get('col_status', self.lang)}</th>
                        <th>{T.get('col_notes', self.lang)}</th>
                    </tr>
            """
            
            for idx, row in enumerate(rows, 1):
                model, kolicina, dodeljeno, status, napomena = row
                slobodno = kolicina - dodeljeno
                
                # Status translation
                status_map = {
                    'Aktivan': T.get('status_active', self.lang),
                    'Na servisu': T.get('status_in_service', self.lang),
                    'Za rashod': T.get('status_for_disposal', self.lang)
                }
                display_status = status_map.get(status, status)
                
                # Color for Slobodno
                slobodno_class = ''
                if slobodno > 0:
                    slobodno_class = 'class="green"'
                elif slobodno == 0:
                    slobodno_class = 'class="red"'
                
                # Color for Status
                status_class = ''
                if status == "Na servisu":
                    status_class = 'class="yellow"'
                elif status == "Za rashod":
                    status_class = 'class="red"'
                
                html += f"""
                    <tr>
                        <td>{idx}</td>
                        <td>{model}</td>
                        <td>{kolicina}</td>
                        <td>{dodeljeno}</td>
                        <td {slobodno_class}>{slobodno}</td>
                        <td {status_class}>{display_status}</td>
                        <td style="text-align: left;">{napomena if napomena else '-'}</td>
                    </tr>
                """
            
            html += f"""
                </table>
                <p class="footer">
                    {'Različitih štampača:' if self.lang == 'sr' else 'Different printers:'} {len(rows)} | 
                    {'Ukupno komada:' if self.lang == 'sr' else 'Total pieces:'} {ukupan_broj}
                </p>
            </body>
            </html>
            """
            
            # Sačuvaj u temp fajl i otvori u browseru
            temp_file = tempfile.NamedTemporaryFile(mode='w', delete=False, suffix='.html', encoding='utf-8')
            temp_file.write(html)
            temp_file.close()
            
            QDesktopServices.openUrl(QUrl.fromLocalFile(temp_file.name))
            
        except Exception as e:
            QMessageBox.critical(self, T.get("error", self.lang), f"{T.get('error', self.lang)}:\n{str(e)}")
    
    def export_stampace_excel(self):
        """Eksportuje sve štampače u Excel"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from datetime import datetime
            
            conn = self.db.get_connection()
            cursor = conn.cursor()
            cursor.execute("""
                SELECT 
                    s.model, 
                    COALESCE(s.kolicina, 1) as kolicina,
                    COUNT(DISTINCT rs.radnik_id) as dodeljeno,
                    s.status,
                    s.napomena
                FROM stampaci s
                LEFT JOIN radnik_stampaci rs ON s.id = rs.stampac_id
                GROUP BY s.id
                ORDER BY s.model
            """)
            rows = cursor.fetchall()
            
            # Izračunaj ukupan broj štampača
            cursor.execute("SELECT COALESCE(SUM(kolicina), 0) FROM stampaci")
            ukupan_broj = cursor.fetchone()[0]
            
            conn.close()
            
            if not rows:
                QMessageBox.information(self, T.get("info", self.lang), "Nema štampača u bazi.")
                return
            
            wb = Workbook()
            ws = wb.active
            ws.title = 'Lista štampača'[:31]
            
            # Naslov
            ws['A1'] = 'LISTA ŠTAMPAČA' if self.lang == 'sr' else 'PRINTER LIST'
            ws['A1'].font = Font(size=18, bold=True)
            ws['A1'].alignment = Alignment(horizontal='center')
            ws.merge_cells('A1:G1')
            
            # Datum
            ws['A2'] = f"{'Datum:' if self.lang == 'sr' else 'Date:'} {datetime.now().strftime('%d.%m.%Y.')}"
            ws['A2'].font = Font(size=11)
            
            # Headers
            headers = [
                'Br.' if self.lang == 'sr' else 'No.',
                T.get('col_model', self.lang),
                T.get('col_quantity', self.lang),
                T.get('col_assigned', self.lang),
                T.get('col_available', self.lang),
                T.get('col_status', self.lang),
                T.get('col_notes', self.lang)
            ]
            
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col)
                cell.value = header
                cell.font = Font(bold=True, color='FFFFFF', size=12)
                cell.fill = PatternFill(start_color='34495E', end_color='34495E', fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
            
            # Data
            for idx, row_data in enumerate(rows, 1):
                model, kolicina, dodeljeno, status, napomena = row_data
                slobodno = kolicina - dodeljeno
                
                # Status translation
                status_map = {
                    'Aktivan': T.get('status_active', self.lang),
                    'Na servisu': T.get('status_in_service', self.lang),
                    'Za rashod': T.get('status_for_disposal', self.lang)
                }
                display_status = status_map.get(status, status)
                
                # Redni broj
                cell = ws.cell(row=idx+4, column=1)
                cell.value = idx
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border
                
                # Model
                cell = ws.cell(row=idx+4, column=2)
                cell.value = model
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border
                
                # Količina
                cell = ws.cell(row=idx+4, column=3)
                cell.value = kolicina
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border
                
                # Dodeljeno
                cell = ws.cell(row=idx+4, column=4)
                cell.value = dodeljeno
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border
                
                # Slobodno
                cell = ws.cell(row=idx+4, column=5)
                cell.value = slobodno
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border
                if slobodno > 0:
                    cell.fill = PatternFill(start_color='C8E6C9', end_color='C8E6C9', fill_type='solid')
                    cell.font = Font(bold=True, color='2E7D32')
                elif slobodno == 0:
                    cell.fill = PatternFill(start_color='FFCDD2', end_color='FFCDD2', fill_type='solid')
                    cell.font = Font(bold=True, color='C62828')
                
                # Status
                cell = ws.cell(row=idx+4, column=6)
                cell.value = display_status
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border
                if status == "Na servisu":
                    cell.fill = PatternFill(start_color='FFF9C4', end_color='FFF9C4', fill_type='solid')
                    cell.font = Font(bold=True, color='F57F17')
                elif status == "Za rashod":
                    cell.fill = PatternFill(start_color='FFCDD2', end_color='FFCDD2', fill_type='solid')
                    cell.font = Font(bold=True, color='C62828')
                
                # Napomena
                cell = ws.cell(row=idx+4, column=7)
                cell.value = napomena if napomena else '-'
                cell.alignment = Alignment(horizontal='left')
                cell.border = thin_border
            
            # Auto width
            for col_idx in range(1, 8):
                max_length = 0
                for row in ws.iter_rows(min_row=4, min_col=col_idx, max_col=col_idx):
                    for cell in row:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                from openpyxl.utils import get_column_letter
                ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 3
            
            # Footer
            footer_row = len(rows) + 6
            if self.lang == 'sr':
                footer_text = f"Različitih štampača: {len(rows)} | Ukupno komada: {ukupan_broj}"
            else:
                footer_text = f"Different printers: {len(rows)} | Total pieces: {ukupan_broj}"
            ws.cell(row=footer_row, column=1).value = footer_text
            ws.cell(row=footer_row, column=1).font = Font(italic=True, size=11)
            
            # Save dialog
            file_path, _ = QFileDialog.getSaveFileName(
                self,
                T.get("btn_excel_export", self.lang),
                f"Lista_stampaca_{datetime.now().strftime('%Y%m%d')}.xlsx",
                "Excel Files (*.xlsx)"
            )
            
            if file_path:
                wb.save(file_path)
                QMessageBox.information(self, T.get("success", self.lang),
                    f"{'Fajl je sačuvan:' if self.lang == 'sr' else 'File saved:'}\n{file_path}")
                
                # Pitaj korisnika da li želi da otvori fajl
                reply = QMessageBox.question(self,
                    T.get("success", self.lang),
                    T.get("msg_open_file", self.lang) if self.lang == 'sr' else "Open file?",
                    QMessageBox.Yes | QMessageBox.No)
                
                if reply == QMessageBox.Yes:
                    QDesktopServices.openUrl(QUrl.fromLocalFile(file_path))
            
        except Exception as e:
            QMessageBox.critical(self, T.get("error", self.lang), f"{T.get('error', self.lang)}:\n{str(e)}")
    
    def preview_narudzbu(self, rows):
        """Prikazuje preview narudžbine u browseru"""
        try:
            from datetime import datetime
            import tempfile
            import os
            
            # Kreiraj HTML sa prevodima
            html = f"""
            <!DOCTYPE html>
            <html>
            <head>
                <meta charset="UTF-8">
                <title>{T.get('order_list_title', self.lang)}</title>
                <style>
                    body {{ 
                        font-family: Arial, sans-serif; 
                        margin: 40px;
                        max-width: 1000px;
                    }}
                    h1 {{ 
                        text-align: center; 
                        color: #2C3E50; 
                        margin-bottom: 10px;
                        font-size: 24pt;
                    }}
                    .datum {{ 
                        text-align: left; 
                        margin-bottom: 30px;
                        font-size: 12pt;
                    }}
                    table {{ 
                        border-collapse: collapse; 
                        width: 100%;
                        box-shadow: 0 2px 4px rgba(0,0,0,0.1);
                    }}
                    th {{ 
                        background-color: #34495E; 
                        color: white; 
                        padding: 15px; 
                        border: 1px solid black;
                        font-size: 14pt;
                    }}
                    td {{ 
                        padding: 12px; 
                        border: 1px solid black; 
                        text-align: center;
                        font-size: 12pt;
                    }}
                    tr:nth-child(even) {{ 
                        background-color: #f2f2f2; 
                    }}
                    tr:hover {{
                        background-color: #e8e8e8;
                    }}
                    .red {{ 
                        background-color: #FFB6B6; 
                        color: #C80000; 
                        font-weight: bold;
                        font-size: 14pt;
                    }}
                    .footer {{ 
                        margin-top: 20px; 
                        font-style: italic;
                        font-size: 11pt;
                    }}
                    .print-button {{
                        background-color: #4CAF50;
                        color: white;
                        padding: 15px 32px;
                        text-align: center;
                        font-size: 16px;
                        margin: 20px 0;
                        cursor: pointer;
                        border: none;
                        border-radius: 4px;
                    }}
                    .print-button:hover {{
                        background-color: #45a049;
                    }}
                    @media print {{
                        .print-button {{ display: none; }}
                        body {{ margin: 20px; }}
                    }}
                </style>
            </head>
            <body>
                <button class="print-button" onclick="window.print()">{T.get('preview_print_btn', self.lang)}</button>
                
                <h1>{T.get('order_list_title', self.lang)}</h1>
                <p class="datum">{T.get('order_date', self.lang)} {datetime.now().strftime('%d.%m.%Y.')}</p>
                <table>
                    <tr>
                        <th>{T.get('col_model', self.lang)}</th>
                        <th>{T.get('col_min_qty', self.lang)}</th>
                        <th>{T.get('col_stock', self.lang)}</th>
                        <th>{T.get('col_for_order', self.lang)}</th>
                    </tr>
            """
            
            for row in rows:
                model, min_kol, trenutno, za_nar = row
                html += f"""
                    <tr>
                        <td>{model}</td>
                        <td>{min_kol}</td>
                        <td>{trenutno}</td>
                        <td class="red">{za_nar}</td>
                    </tr>
                """
            
            html += f"""
                </table>
                <p class="footer">{T.get('order_total', self.lang).format(len(rows))}</p>
            </body>
            </html>
            """
            
            # Sačuvaj u temp fajl i otvori u browseru
            temp_file = tempfile.NamedTemporaryFile(mode='w', delete=False, suffix='.html', encoding='utf-8')
            temp_file.write(html)
            temp_file.close()
            
            QDesktopServices.openUrl(QUrl.fromLocalFile(temp_file.name))
            
        except Exception as e:
            QMessageBox.critical(self, T.get("error", self.lang), f"{T.get('error', self.lang)}:\n{str(e)}")
    
    def stampaj_narudzbu_pdf(self, rows):
        """Štampa narudžbinu direktno na štampač"""
        try:
            from PyQt5.QtPrintSupport import QPrinter, QPrintDialog
            from PyQt5.QtGui import QTextDocument
            from datetime import datetime
            
            # Kreiraj printer
            printer = QPrinter(QPrinter.HighResolution)
            printer.setPageSize(QPrinter.A4)
            
            # Prikaži print dialog
            dialog = QPrintDialog(printer, self)
            if dialog.exec_() != QPrintDialog.Accepted:
                return
            
            # Kreiraj HTML dokument za štampanje
            document = QTextDocument()
            
            # HTML sadržaj
            html = f"""
            <html>
            <head>
                <meta charset="UTF-8">
                <style>
                    body {{ font-family: Arial, sans-serif; margin: 20px; }}
                    h1 {{ text-align: center; color: #2C3E50; margin-bottom: 5px; }}
                    .datum {{ text-align: left; margin-bottom: 20px; }}
                    table {{ border-collapse: collapse; width: 100%; }}
                    th {{ background-color: #34495E; color: white; padding: 10px; border: 1px solid black; }}
                    td {{ padding: 8px; border: 1px solid black; text-align: center; }}
                    tr:nth-child(even) {{ background-color: #f2f2f2; }}
                    .red {{ background-color: #FFB6B6; color: #C80000; font-weight: bold; }}
                    .footer {{ margin-top: 15px; font-style: italic; }}
                </style>
            </head>
            <body>
                <h1>LISTA ZA NARUČIVANJE TONERA</h1>
                <p class="datum">Datum: {datetime.now().strftime('%d.%m.%Y.')}</p>
                <table>
                    <tr>
                        <th>Model</th>
                        <th>Min. količina</th>
                        <th>Trenutno</th>
                        <th>Za naručivanje</th>
                    </tr>
            """
            
            for row in rows:
                model, min_kol, trenutno, za_nar = row
                html += f"""
                    <tr>
                        <td>{model}</td>
                        <td>{min_kol}</td>
                        <td>{trenutno}</td>
                        <td class="red">{za_nar}</td>
                    </tr>
                """
            
            html += f"""
                </table>
                <p class="footer">Ukupno stavki za naručivanje: {len(rows)}</p>
            </body>
            </html>
            """
            
            document.setHtml(html)
            document.print_(printer)
            
            QMessageBox.information(self, T.get("success", self.lang), T.get("msg_sent_to_printer", self.lang))
            
        except Exception as e:
            QMessageBox.critical(self, T.get("error", self.lang), f"Greška prilikom štampanja:\n{str(e)}")
    
    def preview_pregled(self):
        """Prikazuje preview pregleda u browseru"""
        try:
            from datetime import datetime
            import tempfile
            
            conn = self.db.get_connection()
            cursor = conn.cursor()
            
            cursor.execute("""
                SELECT 
                    r.ime || ' ' || r.prezime as radnik,
                    s.model as stampac_model,
                    s.status,
                    GROUP_CONCAT(t.model, ', ') as toneri
                FROM radnici r
                LEFT JOIN radnik_stampaci rs ON r.id = rs.radnik_id
                LEFT JOIN stampaci s ON rs.stampac_id = s.id
                LEFT JOIN stampac_toneri st ON s.id = st.stampac_id
                LEFT JOIN toneri t ON st.toner_id = t.id
                WHERE s.id IS NOT NULL
                GROUP BY r.id, s.id
                ORDER BY r.prezime, r.ime, s.model
            """)
            
            rows = cursor.fetchall()
            conn.close()
            
            if not rows:
                QMessageBox.information(self, T.get("info", self.lang), T.get("error_no_data", self.lang))
                return
            
            # Kreiraj HTML
            html = f"""
            <!DOCTYPE html>
            <html>
            <head>
                <meta charset="UTF-8">
                <title>{T.get('preview_overview_title', self.lang)}</title>
                <style>
                    body {{ 
                        font-family: Arial, sans-serif; 
                        margin: 30px;
                    }}
                    h1 {{ 
                        text-align: center; 
                        color: #2C3E50; 
                        margin-bottom: 10px;
                        font-size: 22pt;
                    }}
                    .datum {{ 
                        text-align: left; 
                        margin-bottom: 25px;
                        font-size: 11pt;
                    }}
                    table {{ 
                        border-collapse: collapse; 
                        width: 100%;
                        box-shadow: 0 2px 4px rgba(0,0,0,0.1);
                    }}
                    th {{ 
                        background-color: #34495E; 
                        color: white; 
                        padding: 12px; 
                        border: 1px solid black;
                        font-size: 12pt;
                    }}
                    td {{ 
                        padding: 10px; 
                        border: 1px solid black;
                        font-size: 11pt;
                    }}
                    tr:nth-child(even) {{ 
                        background-color: #f2f2f2; 
                    }}
                    tr:hover {{
                        background-color: #e8e8e8;
                    }}
                    .footer {{ 
                        margin-top: 15px; 
                        font-style: italic;
                        font-size: 10pt;
                    }}
                    .print-button {{
                        background-color: #4CAF50;
                        color: white;
                        padding: 15px 32px;
                        text-align: center;
                        font-size: 16px;
                        margin: 20px 0;
                        cursor: pointer;
                        border: none;
                        border-radius: 4px;
                    }}
                    .print-button:hover {{
                        background-color: #45a049;
                    }}
                    @media print {{
                        .print-button {{ display: none; }}
                        body {{ margin: 15px; }}
                    }}
                </style>
            </head>
            <body>
                <button class="print-button" onclick="window.print()">{T.get('preview_print_btn', self.lang)}</button>
                
                <h1>{T.get('preview_overview_title', self.lang)}</h1>
                <p class="datum">Datum: {datetime.now().strftime('%d.%m.%Y.')}</p>
                <table>
                    <tr>
                        <th>{T.get('col_employee', self.lang)}</th>
                        <th>{T.get('col_printer', self.lang)}</th>
                        <th>{T.get('col_status', self.lang)}</th>
                        <th>{T.get('col_toners', self.lang)}</th>
                    </tr>
            """
            
            for row in rows:
                radnik, stampac, status, toneri = row
                html += f"""
                    <tr>
                        <td>{radnik if radnik else '-'}</td>
                        <td>{stampac if stampac else '-'}</td>
                        <td>{status if status else '-'}</td>
                        <td>{toneri if toneri else '-'}</td>
                    </tr>
                """
            
            html += f"""
                </table>
                <p class="footer">{T.get('preview_total', self.lang).format(len(rows))}</p>
            </body>
            </html>
            """
            
            # Sačuvaj u temp fajl i otvori u browseru
            temp_file = tempfile.NamedTemporaryFile(mode='w', delete=False, suffix='.html', encoding='utf-8')
            temp_file.write(html)
            temp_file.close()
            
            QDesktopServices.openUrl(QUrl.fromLocalFile(temp_file.name))
                
        except Exception as e:
            QMessageBox.critical(self, T.get("error", self.lang), f"Greška prilikom preview-a:\n{str(e)}")
    
    def stampaj_pregled_pdf(self):
        """Štampa pregled direktno na štampač"""
        try:
            from PyQt5.QtPrintSupport import QPrinter, QPrintDialog
            from PyQt5.QtGui import QTextDocument
            from datetime import datetime
            
            conn = self.db.get_connection()
            cursor = conn.cursor()
            
            cursor.execute("""
                SELECT 
                    r.ime || ' ' || r.prezime as radnik,
                    s.model as stampac_model,
                    s.status,
                    GROUP_CONCAT(t.model, ', ') as toneri
                FROM radnici r
                LEFT JOIN radnik_stampaci rs ON r.id = rs.radnik_id
                LEFT JOIN stampaci s ON rs.stampac_id = s.id
                LEFT JOIN stampac_toneri st ON s.id = st.stampac_id
                LEFT JOIN toneri t ON st.toner_id = t.id
                WHERE s.id IS NOT NULL
                GROUP BY r.id, s.id
                ORDER BY r.prezime, r.ime, s.model
            """)
            
            rows = cursor.fetchall()
            conn.close()
            
            if not rows:
                QMessageBox.information(self, T.get("info", self.lang), T.get("error_no_data", self.lang))
                return
            
            # Kreiraj printer
            printer = QPrinter(QPrinter.HighResolution)
            printer.setPageSize(QPrinter.A4)
            printer.setOrientation(QPrinter.Landscape)
            
            # Prikaži print dialog
            dialog = QPrintDialog(printer, self)
            if dialog.exec_() != QPrintDialog.Accepted:
                return
            
            # Kreiraj HTML dokument
            document = QTextDocument()
            
            html = f"""
            <html>
            <head>
                <meta charset="UTF-8">
                <style>
                    body {{ font-family: Arial, sans-serif; margin: 15px; }}
                    h1 {{ text-align: center; color: #2C3E50; margin-bottom: 5px; font-size: 18pt; }}
                    .datum {{ text-align: left; margin-bottom: 15px; }}
                    table {{ border-collapse: collapse; width: 100%; font-size: 9pt; }}
                    th {{ background-color: #34495E; color: white; padding: 8px; border: 1px solid black; }}
                    td {{ padding: 6px; border: 1px solid black; }}
                    tr:nth-child(even) {{ background-color: #f2f2f2; }}
                    .footer {{ margin-top: 10px; font-style: italic; }}
                </style>
            </head>
            <body>
                <h1>{T.get('preview_overview_title', self.lang)}</h1>
                <p class="datum">Datum: {datetime.now().strftime('%d.%m.%Y.')}</p>
                <table>
                    <tr>
                        <th>{T.get('col_employee', self.lang)}</th>
                        <th>{T.get('col_printer', self.lang)}</th>
                        <th>{T.get('col_status', self.lang)}</th>
                        <th>{T.get('col_toners', self.lang)}</th>
                    </tr>
            """
            
            for row in rows:
                radnik, stampac, status, toneri = row
                html += f"""
                    <tr>
                        <td>{radnik if radnik else '-'}</td>
                        <td>{stampac if stampac else '-'}</td>
                        <td>{status if status else '-'}</td>
                        <td>{toneri if toneri else '-'}</td>
                    </tr>
                """
            
            html += f"""
                </table>
                <p class="footer">{T.get('preview_total', self.lang).format(len(rows))}</p>
            </body>
            </html>
            """
            
            document.setHtml(html)
            document.print_(printer)
            
            QMessageBox.information(self, T.get("success", self.lang), T.get("msg_sent_to_printer", self.lang))
                
        except Exception as e:
            QMessageBox.critical(self, T.get("error", self.lang), f"Greška prilikom štampanja:\n{str(e)}")


    def export_narudzbu_excel(self, rows):
        """Eksportuje narudžbinu u Excel"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from datetime import datetime
            import os
            
            wb = Workbook()
            ws = wb.active
            ws.title = T.get('order_list_title', self.lang)[:31]  # Excel limit 31 chars
            
            # Naslov
            ws['A1'] = T.get('order_list_title', self.lang)
            ws['A1'].font = Font(size=16, bold=True)
            ws['A1'].alignment = Alignment(horizontal='center')
            ws.merge_cells('A1:D1')
            
            # Datum
            ws['A2'] = f"{T.get('order_date', self.lang)} {datetime.now().strftime('%d.%m.%Y.')}"
            ws['A2'].font = Font(size=11)
            
            # Header
            headers = [
                T.get('col_model', self.lang), 
                T.get('col_min_qty', self.lang), 
                T.get('col_stock', self.lang), 
                T.get('col_for_order', self.lang)
            ]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col)
                cell.value = header
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='34495E', end_color='34495E', fill_type='solid')
                cell.alignment = Alignment(horizontal='center')
            
            # Data
            for row_idx, row_data in enumerate(rows, 5):
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.value = value
                    cell.alignment = Alignment(horizontal='center')
                    
                    # Highlight za naručivanje kolona
                    if col_idx == 4:
                        cell.fill = PatternFill(start_color='FFB6B6', end_color='FFB6B6', fill_type='solid')
                        cell.font = Font(bold=True, color='C80000')
            
            # Auto width
            for col_idx in range(1, 5):  # 4 kolone
                max_length = 0
                for row in ws.iter_rows(min_row=4, min_col=col_idx, max_col=col_idx):
                    for cell in row:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                from openpyxl.utils import get_column_letter
                ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 2
            
            # Footer
            footer_row = len(rows) + 6
            if self.lang == 'sr':
                footer_text = f"Ukupno stavki: {len(rows)}"
            else:
                footer_text = f"Total items: {len(rows)}"
            ws.cell(row=footer_row, column=1).value = footer_text
            ws.cell(row=footer_row, column=1).font = Font(italic=True)
            
            # Save
            filename = f"{'Narudzbina' if self.lang == 'sr' else 'Order'}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = os.path.join(os.getcwd(), filename)
            wb.save(filepath)
            
            QMessageBox.information(self, T.get("success", self.lang), T.get("msg_excel_created", self.lang) + "\n\n" + T.get("msg_file", self.lang) + " " + filename)
            
            # Otvori fajl
            import subprocess
            import platform
            if platform.system() == 'Windows':
                os.startfile(filepath)
            elif platform.system() == 'Darwin':
                subprocess.Popen(['open', filepath])
            else:
                subprocess.Popen(['xdg-open', filepath])
                
        except ImportError:
            QMessageBox.warning(self, T.get("error", self.lang), T.get("error_install_openpyxl", self.lang))
        except Exception as e:
            QMessageBox.critical(self, T.get("error", self.lang), f"Greška prilikom kreiranja Excel-a:\n{str(e)}")
    
    def export_pregled_excel(self):
        """Eksportuje pregled u Excel"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from datetime import datetime
            import os
            
            conn = self.db.get_connection()
            cursor = conn.cursor()
            
            cursor.execute("""
                SELECT 
                    r.ime || ' ' || r.prezime as radnik,
                    s.model as stampac_model,
                    s.status,
                    GROUP_CONCAT(t.model, ', ') as toneri
                FROM radnici r
                LEFT JOIN radnik_stampaci rs ON r.id = rs.radnik_id
                LEFT JOIN stampaci s ON rs.stampac_id = s.id
                LEFT JOIN stampac_toneri st ON s.id = st.stampac_id
                LEFT JOIN toneri t ON st.toner_id = t.id
                WHERE s.id IS NOT NULL
                GROUP BY r.id, s.id
                ORDER BY r.prezime, r.ime, s.model
            """)
            
            rows = cursor.fetchall()
            conn.close()
            
            if not rows:
                QMessageBox.information(self, T.get("info", self.lang), T.get("error_no_export_data", self.lang))
                return
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Pregled"
            
            # Naslov
            ws['A1'] = 'PREGLED ŠTAMPAČA I TONERA PO RADNICIMA'
            ws['A1'].font = Font(size=16, bold=True)
            ws['A1'].alignment = Alignment(horizontal='center')
            ws.merge_cells('A1:D1')
            
            # Datum
            ws['A2'] = f"Datum: {datetime.now().strftime('%d.%m.%Y.')}"
            ws['A2'].font = Font(size=11)
            
            # Header
            headers = ['Radnik', 'Štampač', 'Status', 'Toneri']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col)
                cell.value = header
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='34495E', end_color='34495E', fill_type='solid')
                cell.alignment = Alignment(horizontal='center')
            
            # Data
            for row_idx, row_data in enumerate(rows, 5):
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.value = value if value else '-'
                    cell.alignment = Alignment(horizontal='left' if col_idx in [1, 4] else 'center')
            
            # Auto width
            for col_idx in range(1, 5):  # 4 kolone
                max_length = 0
                for row in ws.iter_rows(min_row=4, min_col=col_idx, max_col=col_idx):
                    for cell in row:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                from openpyxl.utils import get_column_letter
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 50)
            
            # Footer
            footer_row = len(rows) + 6
            ws.cell(row=footer_row, column=1).value = f"{T.get('preview_total', self.lang).format(len(rows))}"
            ws.cell(row=footer_row, column=1).font = Font(italic=True)
            
            # Save
            filename = f"Pregled_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = os.path.join(os.getcwd(), filename)
            wb.save(filepath)
            
            QMessageBox.information(self, T.get("success", self.lang), T.get("msg_excel_created", self.lang) + "\n\n" + T.get("msg_file", self.lang) + " " + filename)
            
            # Otvori fajl
            import subprocess
            import platform
            if platform.system() == 'Windows':
                os.startfile(filepath)
            elif platform.system() == 'Darwin':
                subprocess.Popen(['open', filepath])
            else:
                subprocess.Popen(['xdg-open', filepath])
                
        except ImportError:
            QMessageBox.warning(self, T.get("error", self.lang), T.get("error_install_openpyxl", self.lang))
        except Exception as e:
            QMessageBox.critical(self, T.get("error", self.lang), f"Greška prilikom kreiranja Excel-a:\n{str(e)}")



def main():
    app = QApplication(sys.argv)
    app.setStyle('Fusion')  # Modern izgled
    window = MainWindow()
    window.show()
    sys.exit(app.exec_())


if __name__ == '__main__':
    main()